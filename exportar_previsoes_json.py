from openpyxl import load_workbook
from pathlib import Path
from datetime import datetime, time, date
import json
import unicodedata
import re


# ============================================================
# CONFIGURAÇÃO
# ============================================================

FICHEIRO_EXCEL = Path("mundial_2026_db.xlsx")
SEPARADOR = "Calendário"

PASTA_SITE = Path("footwin-vercel")
FICHEIRO_JSON = PASTA_SITE / "public" / "previsoes.json"


# ============================================================
# NORMALIZAÇÃO DE PAÍSES PARA INGLÊS
# ============================================================

COUNTRY_MAP = {
    "mexico": ("Mexico", "mexico"),
    "africa-do-sul": ("South Africa", "south-africa"),
    "south-africa": ("South Africa", "south-africa"),

    "coreia": ("South Korea", "south-korea"),
    "coreia-do-sul": ("South Korea", "south-korea"),
    "republica-da-coreia": ("South Korea", "south-korea"),
    "south-korea": ("South Korea", "south-korea"),

    "chequia": ("Czechia", "czechia"),
    "republica-checa": ("Czechia", "czechia"),
    "czechia": ("Czechia", "czechia"),

    "canada": ("Canada", "canada"),

    "bosnia": ("Bosnia and Herzegovina", "bosnia-and-herzegovina"),
    "bosnia-e-herzegovina": ("Bosnia and Herzegovina", "bosnia-and-herzegovina"),
    "bosnia-and-herzegovina": ("Bosnia and Herzegovina", "bosnia-and-herzegovina"),

    "catar": ("Qatar", "qatar"),
    "qatar": ("Qatar", "qatar"),

    "suica": ("Switzerland", "switzerland"),
    "suicaa": ("Switzerland", "switzerland"),
    "switzerland": ("Switzerland", "switzerland"),

    "brasil": ("Brazil", "brazil"),
    "brazil": ("Brazil", "brazil"),

    "marrocos": ("Morocco", "morocco"),
    "morocco": ("Morocco", "morocco"),

    "haiti": ("Haiti", "haiti"),

    "escocia": ("Scotland", "scotland"),
    "scotland": ("Scotland", "scotland"),

    "estados-unidos": ("United States", "united-states"),
    "usa": ("United States", "united-states"),
    "eua": ("United States", "united-states"),
    "united-states": ("United States", "united-states"),

    "paraguai": ("Paraguay", "paraguay"),
    "paraguay": ("Paraguay", "paraguay"),

    "australia": ("Australia", "australia"),

    "turquia": ("Turkey", "turkey"),
    "turkey": ("Turkey", "turkey"),

    "alemanha": ("Germany", "germany"),
    "germany": ("Germany", "germany"),

    "curacao": ("Curacao", "curacao"),
    "curacau": ("Curacao", "curacao"),
    "curacaoo": ("Curacao", "curacao"),

    "costa-do-marfim": ("Ivory Coast", "ivory-coast"),
    "ivory-coast": ("Ivory Coast", "ivory-coast"),
    "cote-divoire": ("Ivory Coast", "ivory-coast"),
    "cote-d-ivoire": ("Ivory Coast", "ivory-coast"),

    "equador": ("Ecuador", "ecuador"),
    "ecuador": ("Ecuador", "ecuador"),

    "paises-baixos": ("Netherlands", "netherlands"),
    "pais-baixos": ("Netherlands", "netherlands"),
    "netherlands": ("Netherlands", "netherlands"),
    "holanda": ("Netherlands", "netherlands"),

    "japao": ("Japan", "japan"),
    "japan": ("Japan", "japan"),

    "suecia": ("Sweden", "sweden"),
    "sweden": ("Sweden", "sweden"),

    "tunisia": ("Tunisia", "tunisia"),

    "espanha": ("Spain", "spain"),
    "spain": ("Spain", "spain"),

    "cabo-verde": ("Cape Verde", "cape-verde"),
    "cape-verde": ("Cape Verde", "cape-verde"),

    "belgica": ("Belgium", "belgium"),
    "belgium": ("Belgium", "belgium"),

    "egipto": ("Egypt", "egypt"),
    "egito": ("Egypt", "egypt"),
    "egypt": ("Egypt", "egypt"),

    "arabia-saudita": ("Saudi Arabia", "saudi-arabia"),
    "saudi-arabia": ("Saudi Arabia", "saudi-arabia"),

    "uruguai": ("Uruguay", "uruguay"),
    "uruguay": ("Uruguay", "uruguay"),

    "irao": ("Iran", "iran"),
    "iran": ("Iran", "iran"),

    "nova-zelandia": ("New Zealand", "new-zealand"),
    "new-zealand": ("New Zealand", "new-zealand"),

    "franca": ("France", "france"),
    "france": ("France", "france"),

    "senegal": ("Senegal", "senegal"),

    "iraque": ("Iraq", "iraq"),
    "iraq": ("Iraq", "iraq"),

    "noruega": ("Norway", "norway"),
    "norway": ("Norway", "norway"),

    "argentina": ("Argentina", "argentina"),

    "argelia": ("Algeria", "algeria"),
    "algeria": ("Algeria", "algeria"),

    "austria": ("Austria", "austria"),

    "jordania": ("Jordan", "jordan"),
    "jordan": ("Jordan", "jordan"),

    "portugal": ("Portugal", "portugal"),

    "congo-dr": ("DR Congo", "dr-congo"),
    "rd-congo": ("DR Congo", "dr-congo"),
    "dr-congo": ("DR Congo", "dr-congo"),
    "republica-democratica-do-congo": ("DR Congo", "dr-congo"),

    "uzbequistao": ("Uzbekistan", "uzbekistan"),
    "usbequistao": ("Uzbekistan", "uzbekistan"),
    "uzbekistan": ("Uzbekistan", "uzbekistan"),

    "colombia": ("Colombia", "colombia"),

    "inglaterra": ("England", "england"),
    "england": ("England", "england"),

    "croacia": ("Croatia", "croatia"),
    "croatia": ("Croatia", "croatia"),

    "ghana": ("Ghana", "ghana"),
    "gana": ("Ghana", "ghana"),

    "panama": ("Panama", "panama"),
}


# ============================================================
# FUNÇÕES AUXILIARES
# ============================================================

def normalizar_chave(valor):
    if valor is None:
        return ""

    texto = str(valor).strip().lower()
    texto = unicodedata.normalize("NFD", texto)
    texto = "".join(ch for ch in texto if unicodedata.category(ch) != "Mn")
    texto = texto.replace("ç", "c")
    texto = re.sub(r"[^a-z0-9]+", "-", texto)
    texto = re.sub(r"^-+|-+$", "", texto)
    return texto


def pais_ingles(valor):
    chave = normalizar_chave(valor)
    return COUNTRY_MAP.get(chave, (str(valor).strip() if valor is not None else "", chave))[0]


def pais_slug(valor):
    chave = normalizar_chave(valor)
    return COUNTRY_MAP.get(chave, (str(valor).strip() if valor is not None else "", chave))[1]


def valor_limpo(valor):
    if valor is None:
        return ""

    texto = str(valor).strip()

    if texto.lower() in ["none", "nan", "nat"]:
        return ""

    return texto


def formatar_data_excel(valor):
    if valor is None:
        return ""

    if isinstance(valor, datetime):
        return valor.strftime("%Y-%m-%d")

    if isinstance(valor, date):
        return valor.strftime("%Y-%m-%d")

    texto = str(valor).strip()

    if texto.lower() in ["", "none", "nan", "nat"]:
        return ""

    # Se já vier no formato correto
    if re.match(r"^\d{4}-\d{2}-\d{2}$", texto):
        return texto

    # Tenta formatos comuns
    for fmt in ["%d/%m/%Y", "%d-%m-%Y", "%Y/%m/%d", "%Y-%m-%d %H:%M:%S"]:
        try:
            return datetime.strptime(texto, fmt).strftime("%Y-%m-%d")
        except ValueError:
            pass

    return texto


def formatar_hora_excel(valor):
    if valor is None:
        return ""

    if isinstance(valor, datetime):
        return valor.strftime("%H:%M")

    if isinstance(valor, time):
        return valor.strftime("%H:%M")

    # Excel pode guardar horas como fração do dia:
    # 0.0 = 00:00
    # 0.5 = 12:00
    if isinstance(valor, (int, float)):
        total_minutos = round(float(valor) * 24 * 60)
        horas = (total_minutos // 60) % 24
        minutos = total_minutos % 60
        return f"{horas:02d}:{minutos:02d}"

    texto = str(valor).strip()

    if texto.lower() in ["", "none", "nan", "nat"]:
        return ""

    # Corrigir meia-noite
    if texto in ["0", "0.0", "00", "00:00:00"]:
        return "00:00"

    # Se vier com segundos
    if re.match(r"^\d{1,2}:\d{2}:\d{2}$", texto):
        partes = texto.split(":")
        return f"{int(partes[0]):02d}:{int(partes[1]):02d}"

    # Se vier HH:MM ou H:MM
    if re.match(r"^\d{1,2}:\d{2}$", texto):
        partes = texto.split(":")
        return f"{int(partes[0]):02d}:{int(partes[1]):02d}"

    # Se vier tipo 5, assume 05:00
    if texto.isdigit():
        return f"{int(texto):02d}:00"

    return texto


def normalizar_numero(valor):
    texto = valor_limpo(valor)

    if texto == "":
        return ""

    try:
        numero = float(texto)
        if numero.is_integer():
            return str(int(numero))
        return str(numero)
    except ValueError:
        return texto


def tem_resultado_real(jogo):
    return jogo["resultado_real_casa"] != "" and jogo["resultado_real_fora"] != ""


def traduzir_prognostico(prognostico):
    texto = valor_limpo(prognostico)

    if texto == "":
        return ""

    chave = normalizar_chave(texto)

    if chave == "empate" or chave == "draw":
        return "Draw"

    # Formato português: Vitória México
    match_pt = re.match(r"^vit[oó]ria\s+(.+)$", texto, flags=re.IGNORECASE)
    if match_pt:
        equipa = pais_ingles(match_pt.group(1))
        return f"{equipa} win"

    # Formato inglês já correto: Mexico win
    match_en = re.match(r"^(.+)\s+win$", texto, flags=re.IGNORECASE)
    if match_en:
        equipa = pais_ingles(match_en.group(1))
        return f"{equipa} win"

    return texto


def encontrar_colunas(ws):
    colunas = {}

    for cell in ws[1]:
        if cell.value is not None:
            colunas[str(cell.value).strip()] = cell.column

    return colunas


def obter(ws, row, colunas, nome, default=""):
    col = colunas.get(nome)

    if not col:
        return default

    return ws.cell(row=row, column=col).value


def obter_primeira_coluna(ws, row, colunas, nomes, default=""):
    for nome in nomes:
        if nome in colunas:
            return obter(ws, row, colunas, nome, default)

    return default


# ============================================================
# CORREÇÕES MANUAIS DE HORAS CONHECIDAS
# ============================================================

HORAS_CORRIGIDAS = {
    ("Australia", "Turkey"): ("2026-06-14", "05:00"),
    ("Ivory Coast", "Ecuador"): ("2026-06-15", "00:00"),
}


def aplicar_correcoes_manuais(jogo):
    chave = (jogo["casa"], jogo["fora"])

    if chave in HORAS_CORRIGIDAS:
        data_corrigida, hora_corrigida = HORAS_CORRIGIDAS[chave]

        if jogo["data"] == "":
            jogo["data"] = data_corrigida

        if jogo["hora"] == "":
            jogo["hora"] = hora_corrigida

    return jogo


# ============================================================
# EXPORTAÇÃO
# ============================================================

def exportar():
    if not FICHEIRO_EXCEL.exists():
        raise FileNotFoundError(f"Não encontrei o ficheiro Excel: {FICHEIRO_EXCEL}")

    wb = load_workbook(FICHEIRO_EXCEL, data_only=True)

    if SEPARADOR not in wb.sheetnames:
        raise ValueError(f"Não encontrei o separador '{SEPARADOR}' no Excel.")

    ws = wb[SEPARADOR]
    colunas = encontrar_colunas(ws)

    jogos = []

    for row in range(2, ws.max_row + 1):
        casa_original = obter_primeira_coluna(ws, row, colunas, ["Casa", "Equipa Casa", "Home"])
        fora_original = obter_primeira_coluna(ws, row, colunas, ["Fora", "Equipa Fora", "Away"])

        casa_original = valor_limpo(casa_original)
        fora_original = valor_limpo(fora_original)

        if casa_original == "" or fora_original == "":
            continue

        casa = pais_ingles(casa_original)
        fora = pais_ingles(fora_original)

        valor_data_original = obter_primeira_coluna(ws, row, colunas, ["Data", "Date"])

        data = formatar_data_excel(valor_data_original)

        hora = formatar_hora_excel(
            obter_primeira_coluna(ws, row, colunas, ["Hora", "Time"])
        )

        # Se não existir coluna Hora, tenta retirar a hora da própria coluna Data
        if hora == "":
            hora = formatar_hora_excel(valor_data_original)

        grupo = valor_limpo(
            obter_primeira_coluna(ws, row, colunas, ["Grupo", "Group"])
        )

        prognostico = traduzir_prognostico(
            obter_primeira_coluna(ws, row, colunas, ["Prognóstico", "Prognostico", "Prediction"])
        )

        previsao_casa = normalizar_numero(
            obter_primeira_coluna(
                ws,
                row,
                colunas,
                ["Resultado Casa", "Resultado Casa previsto", "Resultado Casa Previsto", "Previsão Casa", "Previsao Casa", "H"],
            )
        )

        previsao_fora = normalizar_numero(
            obter_primeira_coluna(
                ws,
                row,
                colunas,
                ["Resultado Fora", "Resultado Fora previsto", "Resultado Fora Previsto", "Previsão Fora", "Previsao Fora", "I"],
            )
        )

        resultado_real_casa = normalizar_numero(
            obter_primeira_coluna(ws, row, colunas, ["Res C", "Resultado Casa", "Resultado Real Casa"])
        )

        resultado_real_fora = normalizar_numero(
            obter_primeira_coluna(ws, row, colunas, ["Res F", "Resultado Fora", "Resultado Real Fora"])
        )

        jogo = {
            "data": data,
            "hora": hora,
            "grupo": grupo,
            "casa": casa,
            "casa_slug": pais_slug(casa),
            "fora": fora,
            "fora_slug": pais_slug(fora),
            "prognostico": prognostico,
            "previsao_casa": previsao_casa,
            "previsao_fora": previsao_fora,
            "resultado_real_casa": resultado_real_casa,
            "resultado_real_fora": resultado_real_fora,
            "estado": "final" if resultado_real_casa != "" and resultado_real_fora != "" else "previsto",
        }

        jogo = aplicar_correcoes_manuais(jogo)

        jogos.append(jogo)

    # Ordenar por data + hora, mantendo vazios no fim do dia
    def chave_ordenacao(j):
        data = j.get("data") or "9999-12-31"
        hora = j.get("hora") or "99:99"
        return (data, hora, j.get("grupo", ""), j.get("casa", ""))

    jogos.sort(key=chave_ordenacao)

    saida = {
        "atualizado_em": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "total_jogos": len(jogos),
        "jogos": jogos,
    }

    FICHEIRO_JSON.parent.mkdir(parents=True, exist_ok=True)

    FICHEIRO_JSON.write_text(
        json.dumps(saida, ensure_ascii=False, indent=2),
        encoding="utf-8"
    )

    print(f"Exportado com sucesso: {FICHEIRO_JSON}")
    print(f"Total de jogos: {len(jogos)}")

    sem_hora = [j for j in jogos if not j.get("hora")]

    if sem_hora:
        print("\nATENÇÃO: ainda há jogos sem hora:")
        for j in sem_hora:
            print(f"- {j['data']} | {j['casa']} vs {j['fora']}")
    else:
        print("Todos os jogos têm hora.")


if __name__ == "__main__":
    exportar()
