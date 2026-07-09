import json
import re
import hashlib
import html
from datetime import datetime, timedelta
from pathlib import Path
from typing import Any, Dict, List, Optional
from zoneinfo import ZoneInfo

from dateutil import parser as date_parser
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from playwright.sync_api import sync_playwright


# ============================================================
# CONFIGURAÇÕES
# ============================================================

FIFA_URL = (
    "https://www.fifa.com/en/tournaments/mens/worldcup/"
    "canadamexicousa2026/scores-fixtures?country=PT&wtw-filter=ALL"
)

EXCEL_PATH = Path(
    "/Users/admin/Library/Mobile Documents/com~apple~CloudDocs/Footwin World Cup 2026/mundial_2026_db.xlsx"
)

SHEET_NAME = "atualização"

# Tudo será guardado em hora de Portugal continental.
TIMEZONE_LOCAL = ZoneInfo("Europe/Lisbon")

# Só atualizar resultado 2h15 depois da hora de início do jogo.
TEMPO_ESPERA_RESULTADO = timedelta(hours=2, minutes=15)

HEADERS = [
    "game_id",
    "data_hora",
    "atualizacao_programada",
    "fase",
    "grupo",
    "equipa_a",
    "sigla_a",
    "golos_a",
    "equipa_b",
    "sigla_b",
    "golos_b",
    "estado",
    "resultado",
    "vencedor",
    "atualizado_pos_jogo",
    "fonte",
    "ultima_atualizacao",
]


# ============================================================
# FUNÇÕES AUXILIARES
# ============================================================

def normalize_text(value: Any) -> str:
    if value is None:
        return ""

    value = str(value).strip()
    value = re.sub(r"\s+", " ", value)
    return value


def first_value(data: Dict[str, Any], keys: List[str]) -> Any:
    if not isinstance(data, dict):
        return None

    for key in keys:
        if key in data and data[key] not in [None, ""]:
            return data[key]

    return None


def extract_text_from_possible_dict(value: Any) -> str:
    if value is None:
        return ""

    if isinstance(value, str):
        return normalize_text(value)

    if isinstance(value, (int, float)):
        return str(value)

    if isinstance(value, dict):
        possible_keys = [
            "Name",
            "name",
            "ShortName",
            "shortName",
            "DisplayName",
            "displayName",
            "TeamName",
            "teamName",
            "CountryName",
            "countryName",
            "Description",
            "description",
            "Abbreviation",
            "abbreviation",
            "Code",
            "code",
            "TLA",
            "tla",
        ]

        for key in possible_keys:
            if key in value and value[key] not in [None, ""]:
                return extract_text_from_possible_dict(value[key])

        for child in value.values():
            result = extract_text_from_possible_dict(child)
            if result:
                return result

    if isinstance(value, list):
        for item in value:
            result = extract_text_from_possible_dict(item)
            if result:
                return result

    return ""


# ============================================================
# EXTRAÇÃO DE DADOS DOS JOGOS
# ============================================================

def extract_team(match: Dict[str, Any], side: str) -> Dict[str, str]:
    if side == "home":
        team_keys = [
            "HomeTeam",
            "homeTeam",
            "home_team",
            "TeamHome",
            "teamHome",
            "Home",
            "home",
            "HomeContestant",
            "homeContestant",
        ]
        code_keys = [
            "HomeTeamCode",
            "homeTeamCode",
            "HomeTeamAbbreviation",
            "homeTeamAbbreviation",
            "homeCode",
            "HomeTeamShortCode",
        ]
    else:
        team_keys = [
            "AwayTeam",
            "awayTeam",
            "away_team",
            "TeamAway",
            "teamAway",
            "Away",
            "away",
            "AwayContestant",
            "awayContestant",
        ]
        code_keys = [
            "AwayTeamCode",
            "awayTeamCode",
            "AwayTeamAbbreviation",
            "awayTeamAbbreviation",
            "awayCode",
            "AwayTeamShortCode",
        ]

    team_obj = first_value(match, team_keys)
    team_name = extract_text_from_possible_dict(team_obj)

    code = first_value(match, code_keys)
    code = extract_text_from_possible_dict(code)

    if not code and isinstance(team_obj, dict):
        code = extract_text_from_possible_dict(
            first_value(
                team_obj,
                [
                    "Abbreviation",
                    "abbreviation",
                    "Code",
                    "code",
                    "ShortCode",
                    "shortCode",
                    "TLA",
                    "tla",
                ],
            )
        )

    return {
        "name": normalize_text(team_name),
        "code": normalize_text(code),
    }


def extract_score(match: Dict[str, Any], side: str) -> Optional[int]:
    if side == "home":
        keys = [
            "HomeTeamScore",
            "homeTeamScore",
            "HomeScore",
            "homeScore",
            "home_score",
            "ScoreHome",
            "scoreHome",
            "homeGoals",
            "HomeGoals",
        ]
    else:
        keys = [
            "AwayTeamScore",
            "awayTeamScore",
            "AwayScore",
            "awayScore",
            "away_score",
            "ScoreAway",
            "scoreAway",
            "awayGoals",
            "AwayGoals",
        ]

    value = first_value(match, keys)

    if isinstance(value, dict):
        value = first_value(
            value,
            [
                "Score",
                "score",
                "Goals",
                "goals",
                "RegularTime",
                "regularTime",
                "Current",
                "current",
            ],
        )

    if value in [None, ""]:
        return None

    try:
        return int(value)
    except Exception:
        text = str(value)
        found = re.search(r"\d+", text)
        if found:
            return int(found.group())

    return None


def converter_para_hora_portugal(value: Any) -> str:
    """
    A FIFA pode devolver datas com timezone UTC.
    Este método converte sempre para Europe/Lisbon.
    Se a data vier sem timezone, mantém a hora recebida.
    """
    if not value:
        return ""

    try:
        dt = date_parser.parse(str(value))

        if dt.tzinfo is not None:
            dt = dt.astimezone(TIMEZONE_LOCAL).replace(tzinfo=None)

        return dt.strftime("%Y-%m-%d %H:%M")
    except Exception:
        return normalize_text(value)


def extract_date(match: Dict[str, Any]) -> str:
    value = first_value(
        match,
        [
            "Date",
            "date",
            "MatchDate",
            "matchDate",
            "LocalDate",
            "localDate",
            "KickOff",
            "kickOff",
            "Kickoff",
            "kickoff",
            "StartDate",
            "startDate",
            "StartTime",
            "startTime",
        ],
    )

    return converter_para_hora_portugal(value)


def extract_stage(match: Dict[str, Any]) -> str:
    value = first_value(
        match,
        [
            "StageName",
            "stageName",
            "Stage",
            "stage",
            "CompetitionStage",
            "competitionStage",
            "Round",
            "round",
            "Phase",
            "phase",
        ],
    )

    return extract_text_from_possible_dict(value)


def extract_group(match: Dict[str, Any]) -> str:
    value = first_value(
        match,
        [
            "GroupName",
            "groupName",
            "Group",
            "group",
            "Pool",
            "pool",
        ],
    )

    return extract_text_from_possible_dict(value)


def extract_status(match: Dict[str, Any]) -> str:
    value = first_value(
        match,
        [
            "MatchStatus",
            "matchStatus",
            "Status",
            "status",
            "MatchState",
            "matchState",
            "State",
            "state",
        ],
    )

    return extract_text_from_possible_dict(value)


def extract_game_id(
    match: Dict[str, Any],
    equipa_a: str,
    equipa_b: str,
    data_hora: str,
) -> str:
    value = first_value(
        match,
        [
            "IdMatch",
            "idMatch",
            "MatchId",
            "matchId",
            "MatchID",
            "matchID",
            "Id",
            "id",
            "GameId",
            "gameId",
            "FixtureId",
            "fixtureId",
        ],
    )

    if value not in [None, ""]:
        return normalize_text(value)

    raw = f"{data_hora}|{equipa_a}|{equipa_b}"
    return hashlib.md5(raw.encode("utf-8")).hexdigest()


# ============================================================
# CONTROLO DAS 2H15 APÓS INÍCIO DO JOGO
# ============================================================

def parse_data_hora_to_datetime(data_hora: str) -> Optional[datetime]:
    if not data_hora:
        return None

    try:
        return datetime.strptime(data_hora, "%Y-%m-%d %H:%M")
    except Exception:
        try:
            dt = date_parser.parse(data_hora)
            if dt.tzinfo is not None:
                dt = dt.astimezone(TIMEZONE_LOCAL).replace(tzinfo=None)
            return dt
        except Exception:
            return None


def get_atualizacao_programada(data_hora: str) -> str:
    dt = parse_data_hora_to_datetime(data_hora)

    if not dt:
        return ""

    return (dt + TEMPO_ESPERA_RESULTADO).strftime("%Y-%m-%d %H:%M")


def pode_atualizar_resultado(data_hora: str) -> bool:
    dt = parse_data_hora_to_datetime(data_hora)

    if not dt:
        return False

    hora_permitida = dt + TEMPO_ESPERA_RESULTADO
    agora = datetime.now()

    return agora >= hora_permitida


# ============================================================
# DETEÇÃO DE OBJETOS DE JOGO NO JSON
# ============================================================

def is_match_like(obj: Dict[str, Any]) -> bool:
    if not isinstance(obj, dict):
        return False

    home = extract_team(obj, "home")["name"]
    away = extract_team(obj, "away")["name"]

    if home and away and home.lower() != away.lower():
        return True

    return False


def walk_json_for_matches(obj: Any, found: List[Dict[str, Any]]) -> None:
    if isinstance(obj, dict):
        if is_match_like(obj):
            found.append(obj)

        for value in obj.values():
            walk_json_for_matches(value, found)

    elif isinstance(obj, list):
        for item in obj:
            walk_json_for_matches(item, found)


# ============================================================
# CONVERSÃO DE JOGO PARA LINHA DO EXCEL
# ============================================================

def parse_match(match: Dict[str, Any]) -> Optional[Dict[str, Any]]:
    home = extract_team(match, "home")
    away = extract_team(match, "away")

    equipa_a = home["name"]
    equipa_b = away["name"]

    if not equipa_a or not equipa_b:
        return None

    data_hora = extract_date(match)
    atualizacao_programada = get_atualizacao_programada(data_hora)
    permitido_atualizar_resultado = pode_atualizar_resultado(data_hora)

    fase = extract_stage(match)
    grupo = extract_group(match)
    estado = extract_status(match)

    golos_a = extract_score(match, "home")
    golos_b = extract_score(match, "away")

    game_id = extract_game_id(match, equipa_a, equipa_b, data_hora)

    if not permitido_atualizar_resultado:
        golos_a = ""
        golos_b = ""
        resultado = ""
        vencedor = ""
        atualizado_pos_jogo = "Não"
    else:
        atualizado_pos_jogo = "Sim"

        if golos_a is None or golos_b is None:
            golos_a = ""
            golos_b = ""
            resultado = ""
            vencedor = ""
        else:
            resultado = f"{golos_a}-{golos_b}"

            if golos_a > golos_b:
                vencedor = equipa_a
            elif golos_b > golos_a:
                vencedor = equipa_b
            else:
                vencedor = "Empate"

    return {
        "game_id": game_id,
        "data_hora": data_hora,
        "atualizacao_programada": atualizacao_programada,
        "fase": fase,
        "grupo": grupo,
        "equipa_a": equipa_a,
        "sigla_a": home["code"],
        "golos_a": golos_a,
        "equipa_b": equipa_b,
        "sigla_b": away["code"],
        "golos_b": golos_b,
        "estado": estado,
        "resultado": resultado,
        "vencedor": vencedor,
        "atualizado_pos_jogo": atualizado_pos_jogo,
        "fonte": FIFA_URL,
        "ultima_atualizacao": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
    }


# ============================================================
# RECOLHA DA PÁGINA FIFA
# ============================================================

def fetch_fifa_data() -> List[Dict[str, Any]]:
    all_json_objects: List[Any] = []

    with sync_playwright() as p:
        browser = p.chromium.launch(headless=True)

        page = browser.new_page(
            locale="pt-PT",
            timezone_id="Europe/Lisbon",
            viewport={"width": 1600, "height": 1200},
        )

        def handle_response(response):
            try:
                url = response.url.lower()
                content_type = response.headers.get("content-type", "").lower()

                looks_relevant = (
                    "json" in content_type
                    or "api" in url
                    or "match" in url
                    or "fixture" in url
                    or "calendar" in url
                    or "scores" in url
                )

                if looks_relevant:
                    try:
                        data = response.json()
                        all_json_objects.append(data)
                    except Exception:
                        pass

            except Exception:
                pass

        page.on("response", handle_response)

        print("A abrir página da FIFA...")
        page.goto(FIFA_URL, wait_until="domcontentloaded", timeout=120000)

        for label in [
            "Accept All",
            "Accept all",
            "Aceitar tudo",
            "Concordo",
            "Allow all",
            "Permitir tudo",
        ]:
            try:
                page.get_by_text(label, exact=False).click(timeout=2000)
                break
            except Exception:
                pass

        for _ in range(12):
            page.mouse.wheel(0, 2500)
            page.wait_for_timeout(1000)

        html_content = page.content()

        scripts = re.findall(
            r'<script[^>]*type="application/json"[^>]*>(.*?)</script>',
            html_content,
            flags=re.DOTALL | re.IGNORECASE,
        )

        scripts += re.findall(
            r'<script[^>]*id="__NEXT_DATA__"[^>]*>(.*?)</script>',
            html_content,
            flags=re.DOTALL | re.IGNORECASE,
        )

        for script in scripts:
            try:
                cleaned = html.unescape(script).strip()
                data = json.loads(cleaned)
                all_json_objects.append(data)
            except Exception:
                pass

        browser.close()

    raw_matches: List[Dict[str, Any]] = []

    for obj in all_json_objects:
        walk_json_for_matches(obj, raw_matches)

    parsed: List[Dict[str, Any]] = []
    seen = set()

    for raw in raw_matches:
        item = parse_match(raw)

        if not item:
            continue

        key = item["game_id"]

        if key in seen:
            continue

        seen.add(key)
        parsed.append(item)

    parsed.sort(
        key=lambda x: (
            x["data_hora"] or "9999-99-99",
            x["equipa_a"],
            x["equipa_b"],
        )
    )

    return parsed


# ============================================================
# EXCEL
# ============================================================

def get_or_create_workbook(path: Path):
    if path.exists():
        return load_workbook(path)

    wb = Workbook()
    default_sheet = wb.active
    default_sheet.title = "Base"
    return wb


def prepare_sheet(wb):
    if SHEET_NAME in wb.sheetnames:
        ws = wb[SHEET_NAME]
    else:
        ws = wb.create_sheet(SHEET_NAME)

    if ws.max_row == 1 and all(
        ws.cell(1, col).value is None
        for col in range(1, len(HEADERS) + 1)
    ):
        for col, header in enumerate(HEADERS, start=1):
            ws.cell(row=1, column=col, value=header)

    existing_headers = [
        normalize_text(ws.cell(row=1, column=col).value)
        for col in range(1, ws.max_column + 1)
    ]

    for header in HEADERS:
        if header not in existing_headers:
            ws.cell(row=1, column=ws.max_column + 1, value=header)
            existing_headers.append(header)

    return ws


def header_map(ws) -> Dict[str, int]:
    result = {}

    for col in range(1, ws.max_column + 1):
        header = normalize_text(ws.cell(row=1, column=col).value)

        if header:
            result[header] = col

    return result


def style_sheet(ws):
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)

    thin = Side(style="thin", color="D9E2F3")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border

    widths = {
        "A": 18,
        "B": 18,
        "C": 22,
        "D": 22,
        "E": 14,
        "F": 28,
        "G": 10,
        "H": 10,
        "I": 28,
        "J": 10,
        "K": 10,
        "L": 18,
        "M": 12,
        "N": 28,
        "O": 20,
        "P": 75,
        "Q": 22,
    }

    for col_letter, width in widths.items():
        ws.column_dimensions[col_letter].width = width

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(vertical="center")


def update_excel(matches: List[Dict[str, Any]], path: Path):
    wb = get_or_create_workbook(path)
    ws = prepare_sheet(wb)
    hmap = header_map(ws)

    if "game_id" not in hmap:
        raise RuntimeError("Não foi possível encontrar/criar a coluna game_id.")

    game_id_col = hmap["game_id"]

    existing_rows = {}

    for row in range(2, ws.max_row + 1):
        game_id = normalize_text(ws.cell(row=row, column=game_id_col).value)

        if game_id:
            existing_rows[game_id] = row

    inserted = 0
    updated = 0
    skipped_result_update = 0

    for match in matches:
        game_id = match["game_id"]

        if game_id in existing_rows:
            row = existing_rows[game_id]
            updated += 1
        else:
            row = ws.max_row + 1
            existing_rows[game_id] = row
            inserted += 1

        for header in HEADERS:
            col = hmap[header]

            if match.get("atualizado_pos_jogo") == "Não" and header in [
                "golos_a",
                "golos_b",
                "resultado",
                "vencedor",
                "estado",
            ]:
                valor_existente = ws.cell(row=row, column=col).value

                if valor_existente not in [None, ""]:
                    skipped_result_update += 1
                    continue

            ws.cell(row=row, column=col, value=match.get(header, ""))

    style_sheet(ws)

    wb.save(path)

    return inserted, updated, skipped_result_update



def aplicar_resultados_no_calendario(wb):
    """
    Copia resultados finais do separador 'atualização' para o separador 'Calendário'.
    Atualiza TODAS as linhas duplicadas do mesmo jogo.
    """
    if "atualização" not in wb.sheetnames:
        print("AVISO: separador atualização não encontrado.")
        return 0

    if "Calendário" not in wb.sheetnames:
        print("AVISO: separador Calendário não encontrado.")
        return 0

    ws_upd = wb["atualização"]
    ws_cal = wb["Calendário"]

    h_upd = {
        normalize_text(ws_upd.cell(row=1, column=col).value): col
        for col in range(1, ws_upd.max_column + 1)
        if normalize_text(ws_upd.cell(row=1, column=col).value)
    }

    h_cal = {
        normalize_text(ws_cal.cell(row=1, column=col).value): col
        for col in range(1, ws_cal.max_column + 1)
        if normalize_text(ws_cal.cell(row=1, column=col).value)
    }

    obrigatorias_upd = ["equipa_a", "equipa_b", "golos_a", "golos_b", "estado"]
    obrigatorias_cal = ["Casa", "Fora", "Res C", "Res F"]

    for col in obrigatorias_upd:
        if col not in h_upd:
            print(f"AVISO: coluna {col} não encontrada no separador atualização.")
            return 0

    for col in obrigatorias_cal:
        if col not in h_cal:
            print(f"AVISO: coluna {col} não encontrada no separador Calendário.")
            return 0

    def limpo(v):
        return str(v or "").strip()

    def numero(v):
        if v is None:
            return None
        if isinstance(v, (int, float)):
            return int(v)
        t = str(v).strip()
        if t == "":
            return None
        try:
            return int(float(t))
        except ValueError:
            return None

    def estado_final(v):
        e = limpo(v).lower()
        return e in {
            "complete",
            "completed",
            "final",
            "finished",
            "terminado",
            "concluído",
            "concluido",
            "full-time",
            "full time",
            "ft",
        }

    resultados = []

    for r in range(2, ws_upd.max_row + 1):
        equipa_a = limpo(ws_upd.cell(r, h_upd["equipa_a"]).value)
        equipa_b = limpo(ws_upd.cell(r, h_upd["equipa_b"]).value)
        estado = ws_upd.cell(r, h_upd["estado"]).value
        golos_a = numero(ws_upd.cell(r, h_upd["golos_a"]).value)
        golos_b = numero(ws_upd.cell(r, h_upd["golos_b"]).value)

        if not equipa_a or not equipa_b:
            continue

        if golos_a is None or golos_b is None:
            continue

        if not estado_final(estado):
            continue

        resultados.append((equipa_a, equipa_b, golos_a, golos_b))

    print(f"Resultados finais para aplicar no Calendário: {len(resultados)}")

    alteradas = 0

    for equipa_a, equipa_b, golos_a, golos_b in resultados:
        encontrou = False

        for r in range(2, ws_cal.max_row + 1):
            casa = limpo(ws_cal.cell(r, h_cal["Casa"]).value)
            fora = limpo(ws_cal.cell(r, h_cal["Fora"]).value)

            if casa == equipa_a and fora == equipa_b:
                antes_c = ws_cal.cell(r, h_cal["Res C"]).value
                antes_f = ws_cal.cell(r, h_cal["Res F"]).value

                ws_cal.cell(r, h_cal["Res C"]).value = golos_a
                ws_cal.cell(r, h_cal["Res F"]).value = golos_b

                print(
                    f"Calendário linha {r}: {casa} vs {fora} | "
                    f"{antes_c}-{antes_f} -> {golos_a}-{golos_b}"
                )

                alteradas += 1
                encontrou = True

        if not encontrou:
            print(f"AVISO: resultado final não encontrado no Calendário: {equipa_a} vs {equipa_b}")

    return alteradas


# ============================================================
# EXECUÇÃO
# ============================================================

def main():
    print("==============================================")
    print("Atualizador FIFA Mundial 2026")
    print("==============================================")
    print(f"Fonte: {FIFA_URL}")
    print(f"Ficheiro Excel: {EXCEL_PATH}")
    print(f"Separador: {SHEET_NAME}")
    print("Timezone usado: Europe/Lisbon")
    print("Regra: resultados só entram 2h15 após início do jogo.")
    print("==============================================")

    matches = fetch_fifa_data()

    if not matches:
        raise RuntimeError(
            "Não consegui extrair jogos da página da FIFA. "
            "Experimenta correr novamente ou abrir o browser em modo não-headless."
        )

    print(f"Jogos encontrados: {len(matches)}")

    inserted, updated, skipped = update_excel(matches, EXCEL_PATH)

    # Depois de atualizar o separador 'atualização',
    # aplicar automaticamente resultados finais ao separador 'Calendário'
    wb = load_workbook(EXCEL_PATH)
    linhas_calendario = aplicar_resultados_no_calendario(wb)
    wb.save(EXCEL_PATH)
    wb.close()

    print(f"Linhas do Calendário atualizadas com resultados finais: {linhas_calendario}")

    print("Atualização concluída.")
    print(f"Jogos novos adicionados: {inserted}")
    print(f"Jogos atualizados: {updated}")
    print(f"Campos de resultado preservados por ainda não terem passado 2h15: {skipped}")
    print("==============================================")


if __name__ == "__main__":
    main()
