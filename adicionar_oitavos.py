from openpyxl import load_workbook
from pathlib import Path
from datetime import datetime
import shutil
import unicodedata

FICHEIRO = Path("mundial_2026_db.xlsx")
SHEET = "Calendário"

DATA_COL = 1
FASE_COL = 2
CASA_COL = 3
FORA_COL = 4
REAL_CASA_COL = 5
REAL_FORA_COL = 6
PROG_CASA_COL = 8
PROG_FORA_COL = 9
HORA_COL = 25

CONFIG_OITAVOS = [
    ("2026-07-04", "18:00", ("South Africa", "Canada"), ("Netherlands", "Morocco")),
    ("2026-07-04", "22:00", ("Germany", "Paraguay"), ("France", "Sweden")),
    ("2026-07-05", "21:00", ("Brazil", "Japan"), ("Ivory Coast", "Norway")),
    ("2026-07-06", "01:00", ("Mexico", "Ecuador"), ("England", "DR Congo")),
    ("2026-07-06", "20:00", ("Portugal", "Croatia"), ("Spain", "Austria")),
    ("2026-07-07", "01:00", ("United States", "Bosnia and Herzegovina"), ("Belgium", "Senegal")),
    ("2026-07-07", "17:00", ("Argentina", "Cape Verde"), ("Australia", "Egypt")),
    ("2026-07-07", "21:00", ("Switzerland", "Algeria"), ("Colombia", "Ghana")),
]

OVERRIDES_VENCEDOR = {
    ("germany", "paraguay"): "Paraguay",
    ("paraguay", "germany"): "Paraguay",

    ("netherlands", "morocco"): "Morocco",
    ("morocco", "netherlands"): "Morocco",

    ("australia", "egypt"): "Egypt",
    ("egypt", "australia"): "Egypt",

    ("colombia", "ghana"): "Colombia",
    ("ghana", "colombia"): "Colombia",
}


def norm(v):
    s = str(v or "").strip()
    s = unicodedata.normalize("NFKD", s)
    s = "".join(ch for ch in s if not unicodedata.combining(ch))
    s = s.lower()

    aliases = {
        "usa": "united states",
        "u.s.a.": "united states",
        "united states of america": "united states",
        "bosnia & herzegovina": "bosnia and herzegovina",
        "congo dr": "dr congo",
        "cabo verde": "cape verde",
        "cote d'ivoire": "ivory coast",
        "cote d ivoire": "ivory coast",
        "curaçao": "curacao",
        "turkiye": "turkey",
    }

    s = aliases.get(s, s)
    return " ".join(s.split())


def num(v):
    if v is None:
        return None

    try:
        return int(float(v))
    except (TypeError, ValueError):
        return None


def mesmo_jogo(c1, f1, c2, f2):
    return {norm(c1), norm(f1)} == {norm(c2), norm(f2)}


def vencedor_do_jogo(ws, equipa_a, equipa_b):
    na = norm(equipa_a)
    nb = norm(equipa_b)

    override = (
        OVERRIDES_VENCEDOR.get((na, nb))
        or OVERRIDES_VENCEDOR.get((nb, na))
    )

    for r in range(2, ws.max_row + 1):
        casa = ws.cell(r, CASA_COL).value
        fora = ws.cell(r, FORA_COL).value

        if not mesmo_jogo(casa, fora, equipa_a, equipa_b):
            continue

        rc = num(ws.cell(r, REAL_CASA_COL).value)
        rf = num(ws.cell(r, REAL_FORA_COL).value)

        if rc is not None and rf is not None:
            if rc > rf:
                return casa

            if rf > rc:
                return fora

            if override:
                return override

        if override:
            return override

        return None

    if override:
        return override

    return None


def jogo_ja_existe(ws, fase, casa, fora):
    for r in range(2, ws.max_row + 1):
        fase_existente = str(ws.cell(r, FASE_COL).value or "").strip()

        if fase_existente != fase:
            continue

        casa_existente = ws.cell(r, CASA_COL).value
        fora_existente = ws.cell(r, FORA_COL).value

        if mesmo_jogo(casa_existente, fora_existente, casa, fora):
            return True, r

    return False, None


def main():
    if not FICHEIRO.exists():
        raise FileNotFoundError(f"Não encontrei {FICHEIRO}")

    backup = FICHEIRO.with_name(
        f"{FICHEIRO.stem}_backup_antes_verificar_oitavos_"
        f"{datetime.now().strftime('%Y%m%d_%H%M%S')}"
        f"{FICHEIRO.suffix}"
    )

    shutil.copy2(FICHEIRO, backup)
    print(f"Backup criado: {backup.name}")

    wb = load_workbook(FICHEIRO)
    ws = wb[SHEET]

    criados = 0
    existentes = 0
    pendentes = []

    for data, hora, jogo1, jogo2 in CONFIG_OITAVOS:
        vencedor1 = vencedor_do_jogo(ws, jogo1[0], jogo1[1])
        vencedor2 = vencedor_do_jogo(ws, jogo2[0], jogo2[1])

        if not vencedor1 or not vencedor2:
            pendentes.append(
                (data, hora, jogo1, vencedor1, jogo2, vencedor2)
            )
            continue

        existe, linha_existente = jogo_ja_existe(
            ws,
            "R16",
            vencedor1,
            vencedor2
        )

        if existe:
            existentes += 1
            print(
                f"JÁ EXISTE — linha {linha_existente}: "
                f"{data} {hora} | {vencedor1} vs {vencedor2}"
            )
            continue

        linha = ws.max_row + 1

        ws.cell(linha, DATA_COL).value = data
        ws.cell(linha, FASE_COL).value = "R16"
        ws.cell(linha, CASA_COL).value = vencedor1
        ws.cell(linha, FORA_COL).value = vencedor2
        ws.cell(linha, REAL_CASA_COL).value = None
        ws.cell(linha, REAL_FORA_COL).value = None
        ws.cell(linha, PROG_CASA_COL).value = None
        ws.cell(linha, PROG_FORA_COL).value = None
        ws.cell(linha, HORA_COL).value = hora

        criados += 1

        print(
            f"CRIADO — linha {linha}: "
            f"{data} {hora} | {vencedor1} vs {vencedor2}"
        )

    wb.save(FICHEIRO)
    wb.close()

    print("")
    print("==============================================")
    print("RESUMO DOS OITAVOS")
    print("==============================================")
    print(f"Jogos já existentes e preservados: {existentes}")
    print(f"Jogos novos criados: {criados}")
    print(f"Jogos ainda pendentes: {len(pendentes)}")

    if pendentes:
        print("")
        print("PENDENTES:")
        for data, hora, jogo1, vencedor1, jogo2, vencedor2 in pendentes:
            print(
                f"- {data} {hora}: "
                f"{jogo1} -> {vencedor1} | "
                f"{jogo2} -> {vencedor2}"
            )
    else:
        print("")
        print("OK: todos os 8 oitavos estão resolvidos.")

    print("")
    print("Nenhum jogo existente foi apagado.")
    print("Nenhum resultado foi apagado.")
    print("Nenhum prognóstico foi apagado.")


if __name__ == "__main__":
    main()
