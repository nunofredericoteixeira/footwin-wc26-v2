from openpyxl import load_workbook
from pathlib import Path
import subprocess
import sys
from datetime import datetime

EXCEL = Path("mundial_2026_db.xlsx")
BACKUP = Path(f"mundial_2026_db_backup_antes_prognosticos_meias_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx")

def vazio(v):
    return v is None or str(v).strip() == ""

# Backup
BACKUP.write_bytes(EXCEL.read_bytes())
print(f"Backup criado: {BACKUP}")

wb = load_workbook(EXCEL)
ws = wb["Calendário"]

meias = []

for row in range(2, ws.max_row + 1):
    fase = str(ws.cell(row, 2).value or "").strip().upper()

    if fase != "SF":
        continue

    real_casa = ws.cell(row, 5).value
    real_fora = ws.cell(row, 6).value
    prog_txt = ws.cell(row, 7).value
    prog_casa = ws.cell(row, 8).value
    prog_fora = ws.cell(row, 9).value

    if real_casa is not None and real_fora is not None and vazio(prog_txt) and vazio(prog_casa) and vazio(prog_fora):
        meias.append({
            "row": row,
            "real_casa": real_casa,
            "real_fora": real_fora,
        })

if not meias:
    print("Não há meias-finais com resultado real e prognóstico vazio.")
    sys.exit(0)

print("Meias-finais a preencher:")
for m in meias:
    r = m["row"]
    print(f"- linha {r}: {ws.cell(r,3).value} {m['real_casa']}-{m['real_fora']} {ws.cell(r,4).value}")

# Limpar temporariamente resultados reais das meias
for m in meias:
    r = m["row"]
    ws.cell(r, 5).value = None
    ws.cell(r, 6).value = None

wb.save(EXCEL)

# Gerar prognósticos
for i in range(len(meias) + 2):
    print(f"\nGerar prognóstico tentativa {i+1}:")
    result = subprocess.run(
        [sys.executable, "gerar_proximo_prognostico.py"],
        text=True,
        capture_output=True
    )
    print(result.stdout)
    if result.stderr:
        print(result.stderr)

# Repor resultados reais
wb = load_workbook(EXCEL)
ws = wb["Calendário"]

for m in meias:
    r = m["row"]
    ws.cell(r, 5).value = m["real_casa"]
    ws.cell(r, 6).value = m["real_fora"]

wb.save(EXCEL)

print("\nResultados reais repostos. Estado final das meias:")
for m in meias:
    r = m["row"]
    vals = [ws.cell(r, col).value for col in range(1, 10)]
    print(r, vals)

print("OK: prognósticos das meias preenchidos sem apagar resultados reais.")
