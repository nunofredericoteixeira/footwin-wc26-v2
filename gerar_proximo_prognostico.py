#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""
FOOTWIN WORLD CUP 2026
gerar_proximo_prognostico.py

Modelo v13 — ranking + jogadores + 11 provável + utilização + forma + explosão ofensiva.

VERSÃO COM JANELA MÓVEL:
- Mantém no máximo N prognósticos ativos.
- Exemplo:
    python3 gerar_proximo_prognostico.py --janela-prognosticos 3 --debug

Depois de um jogo terminar:
1. O resultado real fica em Res C / Res F.
2. Esse jogo sai da janela.
3. O script recalcula os próximos 3 jogos ainda sem resultado.
4. Limpa todos os prognósticos futuros fora da janela.

Colunas esperadas no Calendário:
A = Data
B = Grupo
C = Casa
D = Fora
E = Res C              -> resultado real casa
F = Res F              -> resultado real fora
G = Prognóstico        -> texto do prognóstico
H = Resultado Casa     -> golos previstos casa para o site
I = Resultado Fora     -> golos previstos fora para o site
"""

from __future__ import annotations

import argparse
import os
import re
import unicodedata
from datetime import datetime, date, time
from typing import Any, Dict, List, Optional, Tuple

from openpyxl import load_workbook


BASE_DIR = os.path.dirname(os.path.abspath(__file__))
EXCEL_FILE = os.path.join(BASE_DIR, "mundial_2026_db.xlsx")

MODELO_VERSAO = "modelo_v13_janela_movel_3"


PESOS_MODELO = {
    "ranking": 0.22,
    "forca_11": 0.30,
    "nivel_clubes": 0.15,
    "utilizacao": 0.13,
    "forma": 0.12,
    "continuidade": 0.05,
    "contexto": 0.03,
}


CLUBES_ELITE = {
    "real madrid", "barcelona", "manchester city", "arsenal", "liverpool",
    "chelsea", "manchester united", "tottenham", "bayern munich", "borussia dortmund",
    "paris saint-germain", "psg", "inter milan", "internazionale", "ac milan",
    "juventus", "napoli", "atletico madrid", "bayer leverkusen", "rb leipzig",
}

CLUBES_FORTES = {
    "benfica", "porto", "sporting", "ajax", "psv", "feyenoord",
    "roma", "lazio", "atalanta", "sevilla", "real sociedad", "valencia",
    "monaco", "marseille", "lyon", "lille", "braga", "galatasaray",
    "fenerbahce", "besiktas", "celtic", "rangers", "brighton", "newcastle",
    "aston villa", "west ham", "fulham", "wolves", "crystal palace",
}

CLUBES_MEDIOS_EUROPA = {
    "mainz", "freiburg", "stuttgart", "wolfsburg", "hoffenheim", "koln",
    "torino", "bologna", "genoa", "sassuolo", "udinese", "empoli",
    "osasuna", "getafe", "celta", "mallorca", "girona", "betis",
    "toulouse", "rennes", "nantes", "montpellier", "reims", "nice",
}

CLUBES_AMERICAS_MEDIOS = {
    "club america", "tigres", "monterrey", "pachuca", "chivas", "guadalajara",
    "cruz azul", "pumas", "toluca", "santos laguna",
    "palmeiras", "flamengo", "fluminense", "sao paulo", "corinthians",
    "river plate", "boca juniors", "racing club", "independiente",
    "atlanta united", "lafc", "inter miami", "la galaxy", "seattle sounders",
}

CLUBES_GOLFO_ASIA = {
    "al sadd", "al-duhail", "al rayyan", "al gharafa", "al arabi",
    "al hilal", "al nassr", "al ahli", "al ittihad", "persepolis",
    "esteghlal", "sepahan", "tractor", "ulsan", "jeonbuk", "seoul",
    "kawasaki", "urawa", "yokohama", "vissel kobe",
}

LIGA_KEYWORDS = [
    ("premier league", 100),
    ("la liga", 96),
    ("bundesliga", 94),
    ("serie a", 92),
    ("ligue 1", 88),
    ("eredivisie", 82),
    ("liga portugal", 80),
    ("primeira liga", 80),
    ("brasileirao", 78),
    ("brasileirão", 78),
    ("argentina", 76),
    ("mls", 72),
    ("saudi", 70),
    ("arabia", 70),
    ("arábia", 70),
    ("turkey", 70),
    ("turquia", 70),
    ("qatar", 58),
    ("iran", 55),
    ("iraq", 48),
]


ALIASES_EQUIPAS = {
    "usa": "united states",
    "eua": "united states",
    "estados unidos": "united states",
    "united states of america": "united states",

    "turquia": "turkey",
    "turkiye": "turkey",
    "türkiye": "turkey",

    "chequia": "czechia",
    "chéquia": "czechia",
    "republica checa": "czechia",
    "república checa": "czechia",
    "czech republic": "czechia",

    "coreia do sul": "south korea",
    "republica da coreia": "south korea",
    "república da coreia": "south korea",
    "korea republic": "south korea",
    "republic of korea": "south korea",

    "paises baixos": "netherlands",
    "países baixos": "netherlands",
    "holanda": "netherlands",

    "marrocos": "morocco",
    "africa do sul": "south africa",
    "áfrica do sul": "south africa",

    "costa do marfim": "ivory coast",
    "cote d ivoire": "ivory coast",
    "côte d ivoire": "ivory coast",

    "curacao": "curacao",
    "curaçao": "curacao",

    "congo dr": "dr congo",
    "rd congo": "dr congo",
    "congo rd": "dr congo",
    "congo democratic republic": "dr congo",
    "democratic republic of congo": "dr congo",

    "espanha": "spain",
    "cabo verde": "cape verde",

    "belgica": "belgium",
    "bélgica": "belgium",

    "egipto": "egypt",
    "egito": "egypt",

    "arabia saudita": "saudi arabia",
    "arábia saudita": "saudi arabia",
    "saudi": "saudi arabia",

    "uruguai": "uruguay",

    "nova zelandia": "new zealand",
    "nova zelândia": "new zealand",

    "franca": "france",
    "frança": "france",

    "iraque": "iraq",
    "noruega": "norway",

    "argelia": "algeria",
    "argélia": "algeria",

    "austria": "austria",
    "áustria": "austria",

    "jordania": "jordan",
    "jordânia": "jordan",

    "colombia": "colombia",
    "colômbia": "colombia",

    "uzbequistao": "uzbekistan",
    "uzbequistão": "uzbekistan",

    "gana": "ghana",

    "panama": "panama",
    "panamá": "panama",
}


def normalizar_txt(v: Any) -> str:
    if v is None:
        return ""
    s = str(v).strip().lower()
    s = unicodedata.normalize("NFKD", s)
    s = "".join(ch for ch in s if not unicodedata.combining(ch))
    s = re.sub(r"[^\w\s:-]", " ", s)
    s = re.sub(r"\s+", " ", s)
    return s.strip()


def chave_equipa(v: Any) -> str:
    s = normalizar_txt(v)
    return ALIASES_EQUIPAS.get(s, s)


def limpar_nome_equipa(v: Any) -> str:
    s = str(v or "").strip()
    s = re.sub(r"\s+", " ", s)
    return s


def to_float(v: Any, default: float = 0.0) -> float:
    if v is None:
        return default
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip().replace(",", ".")
    s = re.sub(r"[^\d.\-]", "", s)
    if not s:
        return default
    try:
        return float(s)
    except ValueError:
        return default


def parse_date(v: Any) -> Optional[date]:
    if v is None:
        return None
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v

    s = str(v).strip()
    if not s:
        return None

    for fmt in ["%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y", "%Y/%m/%d", "%d.%m.%Y"]:
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            pass

    return None


def parse_time(v: Any) -> Optional[time]:
    if v is None:
        return None
    if isinstance(v, datetime):
        return v.time()
    if isinstance(v, time):
        return v

    s = str(v).strip()
    if not s:
        return None

    for fmt in ["%H:%M", "%H:%M:%S"]:
        try:
            return datetime.strptime(s, fmt).time()
        except ValueError:
            pass

    return None


def header_map(ws) -> Dict[str, int]:
    mapa = {}
    for col in range(1, ws.max_column + 1):
        value = ws.cell(row=1, column=col).value
        key = normalizar_txt(value)
        if key:
            mapa[key] = col
    return mapa


def find_col(mapa: Dict[str, int], candidatos: List[str]) -> Optional[int]:
    candidatos_norm = [normalizar_txt(c) for c in candidatos]

    for c in candidatos_norm:
        if c in mapa:
            return mapa[c]

    for key, col in mapa.items():
        for c in candidatos_norm:
            if c and c in key:
                return col

    return None


def find_col_exact(mapa: Dict[str, int], candidatos: List[str]) -> Optional[int]:
    candidatos_norm = [normalizar_txt(c) for c in candidatos]
    for c in candidatos_norm:
        if c in mapa:
            return mapa[c]
    return None


def get_or_create_col(ws, nome: str) -> int:
    mapa = header_map(ws)
    key = normalizar_txt(nome)
    if key in mapa:
        return mapa[key]

    col = ws.max_column + 1
    ws.cell(row=1, column=col).value = nome
    return col


def clamp(v: float, minimo: float, maximo: float) -> float:
    return max(minimo, min(maximo, v))


def obter_ws_calendario(wb):
    for nome in ["Calendário", "Calendario", "calendario", "DB"]:
        if nome in wb.sheetnames:
            return wb[nome]
    raise RuntimeError('Não encontrei separador "Calendário" / "Calendario" / "DB".')


def identificar_colunas_calendario(ws) -> Dict[str, Optional[int]]:
    mapa = header_map(ws)

    return {
        "data": find_col(mapa, ["data", "date", "dia"]),
        "hora": find_col(mapa, ["hora", "time"]),
        "grupo": find_col(mapa, ["grupo", "group"]),
        "local": find_col(mapa, ["campo", "estadio", "estádio", "local", "venue"]),

        "equipa_a": find_col(mapa, ["casa", "equipa_a", "home", "team_a", "seleção_a", "selecao_a", "equipa 1"]),
        "equipa_b": find_col(mapa, ["fora", "equipa_b", "away", "team_b", "seleção_b", "selecao_b", "equipa 2"]),

        # Resultado real
        "res_c": find_col_exact(mapa, ["res c", "res_c", "resultado casa real", "resultado_casa_real"]),
        "res_f": find_col_exact(mapa, ["res f", "res_f", "resultado fora real", "resultado_fora_real"]),

        # Prognóstico visível
        "prognostico": find_col(mapa, ["prognostico", "prognóstico", "previsao", "previsão", "resultado previsto"]),

        # Campos que o site usa para desenhar os números
        "resultado_casa": find_col_exact(mapa, ["resultado casa", "resultado_casa"]),
        "resultado_fora": find_col_exact(mapa, ["resultado fora", "resultado_fora"]),

        "estado": find_col(mapa, ["estado", "status"]),
    }


def estado_terminado(ws, row: int, col_estado: Optional[int]) -> bool:
    if not col_estado:
        return False

    estado = normalizar_txt(ws.cell(row=row, column=col_estado).value)

    return estado in {
        "final",
        "terminado",
        "complete",
        "completed",
        "fim",
        "ft",
        "full time",
        "fulltime",
        "finalizado",
    }


def tem_resultado_real(ws, row: int, cols: Dict[str, Optional[int]]) -> bool:
    if estado_terminado(ws, row, cols.get("estado")):
        return True

    col_res_c = cols.get("res_c")
    col_res_f = cols.get("res_f")

    if col_res_c and col_res_f:
        v_c = ws.cell(row=row, column=col_res_c).value
        v_f = ws.cell(row=row, column=col_res_f).value
        if v_c not in (None, "") and v_f not in (None, ""):
            return True

    return False


def linha_tem_prognostico(ws, row: int, col_prog: Optional[int]) -> bool:
    if not col_prog:
        return False
    val = ws.cell(row=row, column=col_prog).value
    return val not in (None, "")


def sort_key_linha(ws, row: int, cols: Dict[str, Optional[int]]) -> Tuple[datetime, int]:
    d = parse_date(ws.cell(row=row, column=cols["data"]).value) if cols["data"] else None
    h = parse_time(ws.cell(row=row, column=cols["hora"]).value) if cols["hora"] else None

    if not d:
        return datetime.max, row

    if not h:
        h = time(23, 59)

    return datetime.combine(d, h), row


def carregar_forcas_paises(wb) -> Dict[str, Dict[str, Any]]:
    if "Paises" not in wb.sheetnames:
        raise RuntimeError('Separador "Paises" não encontrado no Excel.')

    ws = wb["Paises"]
    mapa = header_map(ws)

    col_pais = find_col(mapa, ["pais", "país", "team", "equipa", "country", "country_name", "seleção", "selecao"])
    col_rank = find_col(mapa, ["rank", "ranking", "posição", "posicao"])
    col_points = find_col(mapa, ["points", "pontos", "pts"])
    col_forca = find_col(mapa, ["forca", "força", "strength"])

    if not col_pais:
        raise RuntimeError('Não encontrei coluna de país/equipa no separador "Paises".')

    linhas = []
    pontos_lista = []
    ranks_lista = []

    for row in range(2, ws.max_row + 1):
        pais = limpar_nome_equipa(ws.cell(row=row, column=col_pais).value)
        if not pais:
            continue

        pontos = to_float(ws.cell(row=row, column=col_points).value, 0) if col_points else 0
        rank = to_float(ws.cell(row=row, column=col_rank).value, 0) if col_rank else 0
        forca = to_float(ws.cell(row=row, column=col_forca).value, 0) if col_forca else 0

        if pontos:
            pontos_lista.append(pontos)
        if rank:
            ranks_lista.append(rank)

        linhas.append({
            "pais": pais,
            "pontos": pontos,
            "rank": rank,
            "forca_excel": forca,
        })

    min_pts = min(pontos_lista) if pontos_lista else 0
    max_pts = max(pontos_lista) if pontos_lista else 0
    min_rank = min(ranks_lista) if ranks_lista else 1
    max_rank = max(ranks_lista) if ranks_lista else 100

    out = {}

    for item in linhas:
        pais = item["pais"]
        key = chave_equipa(pais)

        if item["forca_excel"] > 0:
            forca = item["forca_excel"]
        elif item["pontos"] > 0 and max_pts > min_pts:
            forca = 45 + ((item["pontos"] - min_pts) / (max_pts - min_pts)) * 50
        elif item["rank"] > 0 and max_rank > min_rank:
            forca = 95 - ((item["rank"] - min_rank) / (max_rank - min_rank)) * 50
        else:
            forca = 65

        out[key] = {
            "pais": pais,
            "forca_ranking": clamp(forca, 35, 98),
            "rank": item["rank"],
            "pontos": item["pontos"],
        }

    return out


def obter_forca_ranking(paises: Dict[str, Dict[str, Any]], equipa: str) -> float:
    key = chave_equipa(equipa)
    if key in paises:
        return float(paises[key]["forca_ranking"])
    return 65.0


def descobrir_sheet_jogadores(wb) -> Optional[str]:
    candidatos = [
        "Jogadores",
        "jogadores",
        "Plantel",
        "Planteis",
        "Minutos",
        "Selecoes",
        "Seleções",
        "Base_Jogadores",
        "DB_Jogadores",
    ]

    for nome in candidatos:
        if nome in wb.sheetnames:
            return nome

    for nome in wb.sheetnames:
        if normalizar_txt(nome) in {"calendario", "calendário", "paises", "atualizacao", "atualização"}:
            continue

        ws = wb[nome]
        mapa = header_map(ws)
        tem_pais = find_col(mapa, ["pais_jogador", "pais", "país", "country", "seleção", "selecao"])
        tem_jogador = find_col(mapa, ["jogador", "player", "nome"])
        tem_pos = find_col(mapa, ["posicao", "posição", "position", "pos"])

        if tem_pais and tem_jogador and tem_pos:
            return nome

    return None


def posicao_grupo(posicao: Any) -> str:
    p = normalizar_txt(posicao)

    if p in {"g", "gr", "gk", "guarda redes", "guarda-redes", "goalkeeper"}:
        return "GK"

    if any(x in p for x in ["def", "centre-back", "center-back", "left-back", "right-back", "lateral"]):
        return "DEF"

    if any(x in p for x in ["mid", "medio", "médio", "volante", "meia", "midfielder"]):
        return "MID"

    if any(x in p for x in ["avanc", "avanç", "forward", "striker", "winger", "extremo", "ponta"]):
        return "ATT"

    if p in {"df", "cb", "lb", "rb"}:
        return "DEF"

    if p in {"mc", "cm", "dm", "am", "mf"}:
        return "MID"

    if p in {"fw", "st", "lw", "rw"}:
        return "ATT"

    return "MID"


def forca_clube(clube: Any) -> float:
    c = normalizar_txt(clube)

    if not c:
        return 55

    for nome in CLUBES_ELITE:
        if nome in c:
            return 98

    for nome in CLUBES_FORTES:
        if nome in c:
            return 86

    for nome in CLUBES_MEDIOS_EUROPA:
        if nome in c:
            return 78

    for nome in CLUBES_AMERICAS_MEDIOS:
        if nome in c:
            return 74

    for nome in CLUBES_GOLFO_ASIA:
        if nome in c:
            return 60

    for key, valor in LIGA_KEYWORDS:
        if key in c:
            return valor

    return 62


def carregar_jogadores(wb, paises: Dict[str, Dict[str, Any]]) -> Dict[str, List[Dict[str, Any]]]:
    sheet = descobrir_sheet_jogadores(wb)

    if not sheet:
        return {}

    ws = wb[sheet]
    mapa = header_map(ws)

    col_pais = find_col(mapa, ["pais_jogador", "pais", "país", "country", "seleção", "selecao"])
    col_jogador = find_col(mapa, ["jogador", "player", "nome"])
    col_pos = find_col(mapa, ["posicao", "posição", "position", "pos"])
    col_clube = find_col(mapa, ["clube", "club", "equipa_clube", "time"])

    col_min_poss = find_col(mapa, ["minutos_possiveis", "minutos_possíveis", "minutos_possiveis_selecao", "minutos_possiveis_seleção"])
    col_min_real = find_col(mapa, ["minutos_reais", "minutos_realizados", "minutos"])
    col_jogos_total = find_col(mapa, ["jogos_totais", "jogos_selecao_desde_2025", "jogos_seleção_desde_2025"])
    col_jogos_part = find_col(mapa, ["jogos_participados", "participacoes", "participações"])
    col_util = find_col(mapa, ["percentagem_utilizacao", "percentagem_utilização", "utilizacao", "utilização"])

    if not col_pais or not col_jogador or not col_pos:
        return {}

    out: Dict[str, List[Dict[str, Any]]] = {}

    for row in range(2, ws.max_row + 1):
        pais = limpar_nome_equipa(ws.cell(row=row, column=col_pais).value)
        jogador = str(ws.cell(row=row, column=col_jogador).value or "").strip()
        pos = ws.cell(row=row, column=col_pos).value
        clube = ws.cell(row=row, column=col_clube).value if col_clube else ""

        if not pais or not jogador:
            continue

        key = chave_equipa(pais)

        min_poss = to_float(ws.cell(row=row, column=col_min_poss).value, 0) if col_min_poss else 0
        min_real = to_float(ws.cell(row=row, column=col_min_real).value, 0) if col_min_real else 0
        jogos_total = to_float(ws.cell(row=row, column=col_jogos_total).value, 0) if col_jogos_total else 0
        jogos_part = to_float(ws.cell(row=row, column=col_jogos_part).value, 0) if col_jogos_part else 0

        if col_util:
            util_raw = to_float(ws.cell(row=row, column=col_util).value, 0)
            utilizacao = util_raw / 100 if util_raw > 1 else util_raw
        elif min_poss > 0:
            utilizacao = min_real / min_poss
        elif jogos_total > 0:
            utilizacao = jogos_part / jogos_total
        else:
            utilizacao = 0.35

        utilizacao = clamp(utilizacao, 0, 1)

        ranking_sel = obter_forca_ranking(paises, pais)
        clube_strength = forca_clube(clube)

        fator_utilizacao = 0.50 + utilizacao * 0.50

        forca_j = ((clube_strength * 0.65) + (ranking_sel * 0.35)) * fator_utilizacao
        forca_j = clamp(forca_j, 25, 98)

        item = {
            "pais": pais,
            "jogador": jogador,
            "posicao": posicao_grupo(pos),
            "posicao_original": pos,
            "clube": clube,
            "forca_clube": clube_strength,
            "utilizacao": utilizacao,
            "forca_jogador": forca_j,
        }

        out.setdefault(key, []).append(item)

    return out


def descobrir_sheet_sistemas(wb) -> Optional[str]:
    for nome in ["Sistemas", "Sistema", "Taticas", "Táticas", "Formacoes", "Formações"]:
        if nome in wb.sheetnames:
            return nome
    return None


def carregar_sistemas(wb) -> Dict[str, str]:
    sheet = descobrir_sheet_sistemas(wb)

    if not sheet:
        return {}

    ws = wb[sheet]
    mapa = header_map(ws)

    col_pais = find_col(mapa, ["pais", "país", "country", "seleção", "selecao", "equipa"])
    col_sistema = find_col(mapa, ["sistema", "tatica", "tática", "formacao", "formação"])

    if not col_pais or not col_sistema:
        return {}

    out = {}

    for row in range(2, ws.max_row + 1):
        pais = limpar_nome_equipa(ws.cell(row=row, column=col_pais).value)
        sistema = str(ws.cell(row=row, column=col_sistema).value or "").strip()

        if pais and sistema:
            out[chave_equipa(pais)] = sistema

    return out


def parse_sistema(sistema: str) -> Tuple[int, int, int, int]:
    s = normalizar_txt(sistema)
    nums = [int(x) for x in re.findall(r"\d+", s)]

    if len(nums) == 4:
        return nums[0], nums[1], nums[2], nums[3]

    if len(nums) == 3:
        return 1, nums[0], nums[1], nums[2]

    return 1, 4, 3, 3


def media_top(jogadores: List[Dict[str, Any]], grupo: str, n: int, fallback: float) -> float:
    lista = [j["forca_jogador"] for j in jogadores if j["posicao"] == grupo]
    lista = sorted(lista, reverse=True)

    if not lista:
        return fallback

    if len(lista) < n:
        while len(lista) < n:
            lista.append(fallback * 0.85)

    return sum(lista[:n]) / n


def calcular_forca_11(
    equipa: str,
    jogadores_por_pais: Dict[str, List[Dict[str, Any]]],
    paises: Dict[str, Dict[str, Any]],
    sistemas: Dict[str, str],
) -> Dict[str, float]:
    key = chave_equipa(equipa)
    jogadores = jogadores_por_pais.get(key, [])
    ranking = obter_forca_ranking(paises, equipa)
    sistema = sistemas.get(key, "1-4-3-3")

    gk_n, def_n, mid_n, att_n = parse_sistema(sistema)

    if not jogadores:
        return {
            "forca_11": ranking,
            "nivel_clubes": ranking * 0.90,
            "utilizacao": 0.45,
            "continuidade": 0.45,
            "ataque_11": ranking,
            "defesa_11": ranking,
        }

    gk = media_top(jogadores, "GK", gk_n, ranking)
    defesa = media_top(jogadores, "DEF", def_n, ranking)
    medio = media_top(jogadores, "MID", mid_n, ranking)
    ataque = media_top(jogadores, "ATT", att_n, ranking)

    forca_11 = (
        gk * 0.12
        + defesa * 0.32
        + medio * 0.30
        + ataque * 0.26
    )

    clubes = [j["forca_clube"] for j in jogadores]
    utils = [j["utilizacao"] for j in jogadores]

    nivel_clubes = sum(clubes) / len(clubes) if clubes else ranking * 0.90
    utilizacao = sum(utils) / len(utils) if utils else 0.45

    titulares_habituais = sum(1 for j in jogadores if j["utilizacao"] >= 0.55)
    continuidade = clamp(titulares_habituais / 11, 0, 1)

    return {
        "forca_11": clamp(forca_11, 30, 98),
        "nivel_clubes": clamp(nivel_clubes, 30, 98),
        "utilizacao": clamp(utilizacao, 0, 1),
        "continuidade": clamp(continuidade, 0, 1),
        "ataque_11": clamp(ataque, 30, 98),
        "defesa_11": clamp((gk * 0.35 + defesa * 0.65), 30, 98),
    }


def carregar_historico(wb) -> List[Dict[str, Any]]:
    sheet = None
    for nome in ["Resultados_2025", "Jogos_2025", "Historico", "Histórico", "Resultados", "Forma"]:
        if nome in wb.sheetnames:
            sheet = nome
            break

    if not sheet:
        return []

    ws = wb[sheet]
    mapa = header_map(ws)

    col_data = find_col(mapa, ["data", "date", "dia"])
    col_a = find_col(mapa, ["equipa_a", "casa", "home", "team_a", "seleção_a", "selecao_a"])
    col_b = find_col(mapa, ["equipa_b", "fora", "away", "team_b", "seleção_b", "selecao_b"])
    col_ga = find_col(mapa, ["golos_a", "golosa", "home_score", "score_a", "resultado_a"])
    col_gb = find_col(mapa, ["golos_b", "golosb", "away_score", "score_b", "resultado_b"])

    if not col_a or not col_b or not col_ga or not col_gb:
        return []

    jogos = []

    for row in range(2, ws.max_row + 1):
        equipa_a = limpar_nome_equipa(ws.cell(row=row, column=col_a).value)
        equipa_b = limpar_nome_equipa(ws.cell(row=row, column=col_b).value)

        if not equipa_a or not equipa_b:
            continue

        d = parse_date(ws.cell(row=row, column=col_data).value) if col_data else None
        ga = to_float(ws.cell(row=row, column=col_ga).value, None)
        gb = to_float(ws.cell(row=row, column=col_gb).value, None)

        if ga is None or gb is None:
            continue

        jogos.append({
            "data": d,
            "equipa_a": equipa_a,
            "equipa_b": equipa_b,
            "golos_a": int(ga),
            "golos_b": int(gb),
        })

    return jogos


def calcular_forma_recente(
    equipa: str,
    data_jogo: Optional[date],
    historico: List[Dict[str, Any]],
    paises: Dict[str, Dict[str, Any]],
) -> float:
    key = chave_equipa(equipa)
    jogos_equipa = []

    for j in historico:
        if data_jogo and j["data"] and j["data"] >= data_jogo:
            continue

        a = chave_equipa(j["equipa_a"])
        b = chave_equipa(j["equipa_b"])

        if key == a:
            adv = j["equipa_b"]
            gf = j["golos_a"]
            gs = j["golos_b"]
        elif key == b:
            adv = j["equipa_a"]
            gf = j["golos_b"]
            gs = j["golos_a"]
        else:
            continue

        adv_forca = obter_forca_ranking(paises, adv)
        diff = gf - gs
        fator_adv = 0.85 + (adv_forca / 100) * 0.35

        score = 60 + (diff * 8 * fator_adv) + (gf * 2.5) - (gs * 2.0)
        jogos_equipa.append(score)

    if not jogos_equipa:
        return 65.0

    ultimos = jogos_equipa[-10:]
    return clamp(sum(ultimos) / len(ultimos), 40, 90)


def calcular_contexto(equipa: str, adversario: str, local: Any = None) -> float:
    eq = chave_equipa(equipa)
    loc = chave_equipa(local)

    if eq in {"mexico", "canada", "united states"}:
        return 78

    if loc and eq in loc:
        return 74

    return 50


def calcular_equipa_modelo(
    equipa: str,
    adversario: str,
    data_jogo: Optional[date],
    local: Any,
    paises: Dict[str, Dict[str, Any]],
    jogadores_por_pais: Dict[str, List[Dict[str, Any]]],
    sistemas: Dict[str, str],
    historico: List[Dict[str, Any]],
) -> Dict[str, float]:

    ranking = obter_forca_ranking(paises, equipa)
    xi = calcular_forca_11(equipa, jogadores_por_pais, paises, sistemas)
    forma = calcular_forma_recente(equipa, data_jogo, historico, paises)
    contexto = calcular_contexto(equipa, adversario, local)

    utilizacao_score = 40 + xi["utilizacao"] * 60
    continuidade_score = 40 + xi["continuidade"] * 60

    forca_final = (
        ranking * PESOS_MODELO["ranking"]
        + xi["forca_11"] * PESOS_MODELO["forca_11"]
        + xi["nivel_clubes"] * PESOS_MODELO["nivel_clubes"]
        + utilizacao_score * PESOS_MODELO["utilizacao"]
        + forma * PESOS_MODELO["forma"]
        + continuidade_score * PESOS_MODELO["continuidade"]
        + contexto * PESOS_MODELO["contexto"]
    )

    return {
        "ranking": ranking,
        "forca_11": xi["forca_11"],
        "nivel_clubes": xi["nivel_clubes"],
        "utilizacao_score": utilizacao_score,
        "forma": forma,
        "continuidade_score": continuidade_score,
        "contexto": contexto,
        "forca_final": clamp(forca_final, 35, 98),
        "ataque_11": xi["ataque_11"],
        "defesa_11": xi["defesa_11"],
    }


def calcular_xg(equipa: Dict[str, float], adversario: Dict[str, float]) -> float:
    ataque = equipa["ataque_11"]
    defesa_adv = adversario["defesa_11"]
    diff_forca = equipa["forca_final"] - adversario["forca_final"]

    xg = 1.15
    xg += (ataque - defesa_adv) / 38.0
    xg += diff_forca / 42.0
    xg += (equipa["forma"] - 65) / 55.0
    xg += (equipa["contexto"] - 50) / 80.0
    xg += (equipa["utilizacao_score"] - 65) / 120.0

    return clamp(xg, 0.15, 5.50)


def aplicar_explosao_ofensiva(
    xg_a: float,
    xg_b: float,
    a: Dict[str, float],
    b: Dict[str, float],
) -> Tuple[float, float]:

    def boost(xg_fav: float, fav: Dict[str, float], dog: Dict[str, float]) -> float:
        diff_final = fav["forca_final"] - dog["forca_final"]
        diff_ataque_defesa = fav["ataque_11"] - dog["defesa_11"]
        diff_clubes = fav["nivel_clubes"] - dog["nivel_clubes"]

        indice_explosao = (
            diff_final * 0.45
            + diff_ataque_defesa * 0.40
            + diff_clubes * 0.15
        )

        if indice_explosao >= 18:
            xg_fav += 0.35

        if indice_explosao >= 26:
            xg_fav += 0.70

        if indice_explosao >= 36:
            xg_fav += 1.10

        if fav["ataque_11"] >= 78 and dog["defesa_11"] <= 62:
            xg_fav += 0.45

        return clamp(xg_fav, 0.15, 6.80)

    if a["forca_final"] > b["forca_final"]:
        xg_a = boost(xg_a, a, b)
    elif b["forca_final"] > a["forca_final"]:
        xg_b = boost(xg_b, b, a)

    return xg_a, xg_b


def converter_xg_para_golos(xg: float) -> int:
    if xg < 0.45:
        return 0
    if xg < 1.25:
        return 1
    if xg < 2.05:
        return 2
    if xg < 2.85:
        return 3
    if xg < 3.65:
        return 4
    if xg < 4.55:
        return 5
    if xg < 5.45:
        return 6
    return 7


def ajustar_resultado_final(
    golos_a: int,
    golos_b: int,
    xg_a: float,
    xg_b: float,
    a: Dict[str, float],
    b: Dict[str, float],
) -> Tuple[int, int]:

    diff_forca = a["forca_final"] - b["forca_final"]

    if abs(diff_forca) < 4 and 0.85 <= xg_a <= 1.55 and 0.85 <= xg_b <= 1.55:
        return 1, 1

    if abs(diff_forca) < 6 and xg_a >= 1.65 and xg_b >= 1.55:
        return 2, 2

    if 4 <= diff_forca < 9 and golos_a - golos_b >= 2:
        golos_a = max(golos_b + 1, 1)

    if -9 < diff_forca <= -4 and golos_b - golos_a >= 2:
        golos_b = max(golos_a + 1, 1)

    if diff_forca >= 12 and xg_a >= 1.75:
        golos_a = max(golos_a, 2)

    if diff_forca <= -12 and xg_b >= 1.75:
        golos_b = max(golos_b, 2)

    if xg_b < 0.65 and diff_forca >= 10:
        golos_b = 0

    if xg_a < 0.65 and diff_forca <= -10:
        golos_a = 0

    return int(golos_a), int(golos_b)


def prever_jogo(
    equipa_a: str,
    equipa_b: str,
    data_jogo: Optional[date],
    local: Any,
    paises: Dict[str, Dict[str, Any]],
    jogadores_por_pais: Dict[str, List[Dict[str, Any]]],
    sistemas: Dict[str, str],
    historico: List[Dict[str, Any]],
) -> Dict[str, Any]:

    modelo_a = calcular_equipa_modelo(
        equipa_a, equipa_b, data_jogo, local,
        paises, jogadores_por_pais, sistemas, historico
    )

    modelo_b = calcular_equipa_modelo(
        equipa_b, equipa_a, data_jogo, local,
        paises, jogadores_por_pais, sistemas, historico
    )

    xg_a = calcular_xg(modelo_a, modelo_b)
    xg_b = calcular_xg(modelo_b, modelo_a)

    xg_a, xg_b = aplicar_explosao_ofensiva(xg_a, xg_b, modelo_a, modelo_b)

    golos_a = converter_xg_para_golos(xg_a)
    golos_b = converter_xg_para_golos(xg_b)

    golos_a, golos_b = ajustar_resultado_final(
        golos_a, golos_b, xg_a, xg_b, modelo_a, modelo_b
    )

    prognostico = f"{golos_a}-{golos_b}"

    return {
        "equipa_a": equipa_a,
        "equipa_b": equipa_b,
        "golos_a": golos_a,
        "golos_b": golos_b,
        "prognostico": prognostico,
        "xg_a": round(xg_a, 2),
        "xg_b": round(xg_b, 2),
        "forca_a": round(modelo_a["forca_final"], 2),
        "forca_b": round(modelo_b["forca_final"], 2),
        "ataque_a": round(modelo_a["ataque_11"], 2),
        "ataque_b": round(modelo_b["ataque_11"], 2),
        "defesa_a": round(modelo_a["defesa_11"], 2),
        "defesa_b": round(modelo_b["defesa_11"], 2),
        "modelo": MODELO_VERSAO,
    }


def match_jogo_pedido(equipa_a: str, equipa_b: str, jogos_pedidos: Optional[List[str]]) -> bool:
    if not jogos_pedidos:
        return True

    a = chave_equipa(equipa_a)
    b = chave_equipa(equipa_b)

    for item in jogos_pedidos:
        if not item:
            continue

        partes = item.split(":")

        if len(partes) != 2:
            continue

        j1 = chave_equipa(partes[0])
        j2 = chave_equipa(partes[1])

        if a == j1 and b == j2:
            return True

        if a == j2 and b == j1:
            return True

    return False


def obter_linhas_futuras_sem_resultado(ws, cols: Dict[str, Optional[int]]) -> List[int]:
    if not cols["equipa_a"] or not cols["equipa_b"]:
        raise RuntimeError("Não encontrei colunas das equipas no Calendário.")

    linhas = []

    for row in range(2, ws.max_row + 1):
        equipa_a = limpar_nome_equipa(ws.cell(row=row, column=cols["equipa_a"]).value)
        equipa_b = limpar_nome_equipa(ws.cell(row=row, column=cols["equipa_b"]).value)

        if not equipa_a or not equipa_b:
            continue

        if tem_resultado_real(ws, row, cols):
            continue

        linhas.append(row)

    linhas.sort(key=lambda r: sort_key_linha(ws, r, cols))

    return linhas


def selecionar_linhas_para_prever(
    ws,
    cols: Dict[str, Optional[int]],
    todos_futuros: bool,
    recalcular_futuros: bool,
    jogos_pedidos: Optional[List[str]] = None,
) -> List[int]:

    if not cols["equipa_a"] or not cols["equipa_b"]:
        raise RuntimeError("Não encontrei colunas das equipas no Calendário.")

    linhas = []

    for row in range(2, ws.max_row + 1):
        equipa_a = limpar_nome_equipa(ws.cell(row=row, column=cols["equipa_a"]).value)
        equipa_b = limpar_nome_equipa(ws.cell(row=row, column=cols["equipa_b"]).value)

        if not equipa_a or not equipa_b:
            continue

        if not match_jogo_pedido(equipa_a, equipa_b, jogos_pedidos):
            continue

        if jogos_pedidos:
            if tem_resultado_real(ws, row, cols):
                continue
            linhas.append(row)
            continue

        if tem_resultado_real(ws, row, cols):
            continue

        if linha_tem_prognostico(ws, row, cols["prognostico"]) and not recalcular_futuros:
            continue

        linhas.append(row)

    linhas.sort(key=lambda r: sort_key_linha(ws, r, cols))

    if jogos_pedidos:
        return linhas

    if todos_futuros or recalcular_futuros:
        return linhas

    return linhas[:1]


def escrever_prognostico(ws, row: int, previsao: Dict[str, Any]) -> None:
    """
    Escreve em:
    G = Prognóstico
    H = Resultado Casa
    I = Resultado Fora
    e também em colunas auxiliares.
    """

    prognostico = previsao["prognostico"]
    golos_a = previsao["golos_a"]
    golos_b = previsao["golos_b"]

    col_prog = get_or_create_col(ws, "Prognóstico")
    col_resultado_casa = get_or_create_col(ws, "Resultado Casa")
    col_resultado_fora = get_or_create_col(ws, "Resultado Fora")
    col_modelo = get_or_create_col(ws, "Modelo")
    col_xga = get_or_create_col(ws, "xG_A")
    col_xgb = get_or_create_col(ws, "xG_B")
    col_forca_a = get_or_create_col(ws, "Força_A")
    col_forca_b = get_or_create_col(ws, "Força_B")
    col_obs = get_or_create_col(ws, "Obs_Modelo")

    ws.cell(row=row, column=col_prog).value = prognostico
    ws.cell(row=row, column=col_resultado_casa).value = golos_a
    ws.cell(row=row, column=col_resultado_fora).value = golos_b
    ws.cell(row=row, column=col_modelo).value = previsao["modelo"]
    ws.cell(row=row, column=col_xga).value = previsao["xg_a"]
    ws.cell(row=row, column=col_xgb).value = previsao["xg_b"]
    ws.cell(row=row, column=col_forca_a).value = previsao["forca_a"]
    ws.cell(row=row, column=col_forca_b).value = previsao["forca_b"]

    mapa = header_map(ws)

    nomes_colunas_texto = [
        "prognostico",
        "prognóstico",
        "previsao",
        "previsão",
        "resultado previsto",
        "resultado_previsto",
        "resultado predito",
        "resultado_predito",
        "placar previsto",
        "score previsto",
    ]

    nomes_colunas_golos_a = [
        "resultado casa",
        "resultado_casa",
        "golos_previstos_a",
        "golos previsto a",
        "golos_prognostico_a",
        "golos prognostico a",
        "golos_prognóstico_a",
        "previsao_golos_a",
        "previsão_golos_a",
        "golos casa previsto",
        "golos_casa_previsto",
        "golos equipa a previsto",
        "golos_equipa_a_previsto",
        "home_pred",
        "home_prediction",
        "pred_home",
        "forecast_home",
    ]

    nomes_colunas_golos_b = [
        "resultado fora",
        "resultado_fora",
        "golos_previstos_b",
        "golos previsto b",
        "golos_prognostico_b",
        "golos prognostico b",
        "golos_prognóstico_b",
        "previsao_golos_b",
        "previsão_golos_b",
        "golos fora previsto",
        "golos_fora_previsto",
        "golos equipa b previsto",
        "golos_equipa_b_previsto",
        "away_pred",
        "away_prediction",
        "pred_away",
        "forecast_away",
    ]

    for nome in nomes_colunas_texto:
        key = normalizar_txt(nome)
        if key in mapa:
            ws.cell(row=row, column=mapa[key]).value = prognostico

    for nome in nomes_colunas_golos_a:
        key = normalizar_txt(nome)
        if key in mapa:
            ws.cell(row=row, column=mapa[key]).value = golos_a

    for nome in nomes_colunas_golos_b:
        key = normalizar_txt(nome)
        if key in mapa:
            ws.cell(row=row, column=mapa[key]).value = golos_b

    obs = (
        f"{previsao['equipa_a']} {prognostico} {previsao['equipa_b']} | "
        f"xG {previsao['xg_a']}-{previsao['xg_b']} | "
        f"Forças {previsao['forca_a']}-{previsao['forca_b']} | "
        f"Ataques {previsao['ataque_a']}-{previsao['ataque_b']} | "
        f"Defesas {previsao['defesa_a']}-{previsao['defesa_b']}"
    )

    ws.cell(row=row, column=col_obs).value = obs


def limpar_prognostico(ws, row: int) -> None:
    """
    Limpa prognóstico fora da janela.
    Não limpa Res C / Res F, porque esses são resultado real.
    """

    mapa = header_map(ws)

    nomes_para_limpar = [
        "prognostico",
        "prognóstico",
        "previsao",
        "previsão",
        "resultado previsto",
        "resultado_previsto",
        "resultado predito",
        "resultado_predito",
        "placar previsto",
        "score previsto",

        "resultado casa",
        "resultado_casa",
        "resultado fora",
        "resultado_fora",

        "golos_previstos_a",
        "golos previsto a",
        "golos_prognostico_a",
        "golos prognostico a",
        "golos_prognóstico_a",
        "previsao_golos_a",
        "previsão_golos_a",
        "golos casa previsto",
        "golos_casa_previsto",
        "golos equipa a previsto",
        "golos_equipa_a_previsto",
        "home_pred",
        "home_prediction",
        "pred_home",
        "forecast_home",

        "golos_previstos_b",
        "golos previsto b",
        "golos_prognostico_b",
        "golos prognostico b",
        "golos_prognóstico_b",
        "previsao_golos_b",
        "previsão_golos_b",
        "golos fora previsto",
        "golos_fora_previsto",
        "golos equipa b previsto",
        "golos_equipa_b_previsto",
        "away_pred",
        "away_prediction",
        "pred_away",
        "forecast_away",

        "modelo",
        "xg_a",
        "xg_b",
        "força_a",
        "forca_a",
        "força_b",
        "forca_b",
        "obs_modelo",
    ]

    for nome in nomes_para_limpar:
        key = normalizar_txt(nome)
        if key in mapa:
            ws.cell(row=row, column=mapa[key]).value = None


def listar_jogos_disponiveis(ws, cols: Dict[str, Optional[int]], max_linhas: int = 100) -> None:
    print("Jogos encontrados no calendário:")
    n = 0

    for row in range(2, ws.max_row + 1):
        if not cols["equipa_a"] or not cols["equipa_b"]:
            return

        equipa_a = limpar_nome_equipa(ws.cell(row=row, column=cols["equipa_a"]).value)
        equipa_b = limpar_nome_equipa(ws.cell(row=row, column=cols["equipa_b"]).value)

        if not equipa_a or not equipa_b:
            continue

        res_c = ws.cell(row=row, column=cols["res_c"]).value if cols["res_c"] else ""
        res_f = ws.cell(row=row, column=cols["res_f"]).value if cols["res_f"] else ""
        prog = ws.cell(row=row, column=cols["prognostico"]).value if cols["prognostico"] else ""

        estado = "FINAL" if tem_resultado_real(ws, row, cols) else "POR JOGAR"

        print(f" - Linha {row}: {equipa_a} vs {equipa_b} | real: {res_c}-{res_f} | prog: {prog} | {estado}")

        n += 1
        if n >= max_linhas:
            break


def main():
    parser = argparse.ArgumentParser()

    parser.add_argument("--todos-futuros", action="store_true", help="gera prognóstico para todos os jogos futuros sem prognóstico")
    parser.add_argument("--recalcular-futuros", action="store_true", help="recalcula todos os jogos futuros sem resultado real")
    parser.add_argument("--limite", type=int, default=None, help="limita o número de jogos")
    parser.add_argument("--jogo", action="append", default=None, help='recalcula jogo específico no formato "Equipa A:Equipa B"')
    parser.add_argument("--janela-prognosticos", type=int, default=None, help="mantém apenas N prognósticos ativos")
    parser.add_argument("--listar-jogos", action="store_true", help="lista jogos encontrados no calendário")
    parser.add_argument("--debug", action="store_true", help="mostra detalhes")

    args = parser.parse_args()

    if not os.path.exists(EXCEL_FILE):
        raise FileNotFoundError(f"Excel não encontrado: {EXCEL_FILE}")

    print("==============================================")
    print("FOOTWIN WORLD CUP 2026 — GERAR PROGNÓSTICO")
    print(f"Modelo: {MODELO_VERSAO}")
    print("==============================================")

    wb = load_workbook(EXCEL_FILE)
    ws_cal = obter_ws_calendario(wb)
    cols = identificar_colunas_calendario(ws_cal)

    if not cols["prognostico"]:
        cols["prognostico"] = get_or_create_col(ws_cal, "Prognóstico")
    if not cols["resultado_casa"]:
        cols["resultado_casa"] = get_or_create_col(ws_cal, "Resultado Casa")
    if not cols["resultado_fora"]:
        cols["resultado_fora"] = get_or_create_col(ws_cal, "Resultado Fora")

    if args.listar_jogos:
        listar_jogos_disponiveis(ws_cal, cols)
        return

    paises = carregar_forcas_paises(wb)
    jogadores_por_pais = carregar_jogadores(wb, paises)
    sistemas = carregar_sistemas(wb)
    historico = carregar_historico(wb)

    print(f"Equipas em Paises: {len(paises)}")
    print(f"Seleções com jogadores/minutos: {len(jogadores_por_pais)}")
    print(f"Sistemas táticos carregados: {len(sistemas)}")
    print(f"Jogos históricos carregados: {len(historico)}")
    print("----------------------------------------------")

    if args.janela_prognosticos is not None:
        todas_futuras = obter_linhas_futuras_sem_resultado(ws_cal, cols)

        linhas = todas_futuras[:args.janela_prognosticos]
        linhas_para_limpar = todas_futuras[args.janela_prognosticos:]

        print(f"Modo janela ativa: máximo {args.janela_prognosticos} prognósticos.")
        print(f"Jogos a recalcular: {len(linhas)}")
        print(f"Jogos futuros a limpar: {len(linhas_para_limpar)}")
        print("----------------------------------------------")

        for row_limpar in linhas_para_limpar:
            limpar_prognostico(ws_cal, row_limpar)

    else:
        linhas = selecionar_linhas_para_prever(
            ws_cal,
            cols,
            todos_futuros=args.todos_futuros,
            recalcular_futuros=args.recalcular_futuros,
            jogos_pedidos=args.jogo,
        )

        if args.limite is not None:
            linhas = linhas[:args.limite]

    if not linhas:
        print("Não há jogos elegíveis para prognóstico.")
        return

    for row in linhas:
        equipa_a = limpar_nome_equipa(ws_cal.cell(row=row, column=cols["equipa_a"]).value)
        equipa_b = limpar_nome_equipa(ws_cal.cell(row=row, column=cols["equipa_b"]).value)
        data_jogo = parse_date(ws_cal.cell(row=row, column=cols["data"]).value) if cols["data"] else None
        local = ws_cal.cell(row=row, column=cols["local"]).value if cols["local"] else None

        previsao = prever_jogo(
            equipa_a=equipa_a,
            equipa_b=equipa_b,
            data_jogo=data_jogo,
            local=local,
            paises=paises,
            jogadores_por_pais=jogadores_por_pais,
            sistemas=sistemas,
            historico=historico,
        )

        escrever_prognostico(ws_cal, row, previsao)

        print(f"Linha {row}: {equipa_a} vs {equipa_b}")
        print(f"Prognóstico: {equipa_a} {previsao['prognostico']} {equipa_b}")
        print(f"xG: {previsao['xg_a']} - {previsao['xg_b']}")
        print(f"Forças: {previsao['forca_a']} - {previsao['forca_b']}")

        if args.debug:
            print(f"Ataques: {previsao['ataque_a']} - {previsao['ataque_b']}")
            print(f"Defesas: {previsao['defesa_a']} - {previsao['defesa_b']}")

        print("----------------------------------------------")

    wb.save(EXCEL_FILE)
    print("Excel atualizado com sucesso.")
    print(EXCEL_FILE)


if __name__ == "__main__":
    main()