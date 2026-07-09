from openpyxl import load_workbook
from pathlib import Path
from datetime import datetime
import shutil
import unicodedata

FICHEIRO = Path("mundial_2026_db.xlsx")

SHEET_ATUALIZACAO = "atualização"
SHEET_CALENDARIO = "Calendário"

# Calendário
CAL_DATA_COL = 1
CAL_FASE_COL = 2
CAL_CASA_COL = 3
CAL_FORA_COL = 4
CAL_REAL_CASA_COL = 5
CAL_REAL_FORA_COL = 6

# atualização
ATU_FASE_COL = 5
ATU_CASA_COL = 7
ATU_RESULTADO_CASA_COL = 9
ATU_FORA_COL = 10
ATU_RESULTADO_FORA_COL = 12


def norm(v):
    s = str(v or "").strip()
    s = unicodedata.normalize("NFKD", s)
    s = "".join(
        ch for ch in s
        if not unicodedata.combining(ch)
    )
    s = s.lower()

    aliases = {
        "usa": "united states",
        "u.s.a.": "united states",
        "united states of america": "united states",

        "korea republic": "south korea",

        "congo dr": "dr congo",
        "democratic republic of the congo": "dr congo",

        "cabo verde": "cape verde",

        "cote d'ivoire": "ivory coast",
        "cote d ivoire": "ivory coast",

        "turkiye": "turkey",

        "ir iran": "iran",

        "curacao": "curacao",
    }

    s = aliases.get(s, s)

    return " ".join(s.split())


def numero(v):
    if v is None:
        return None

    try:
        return int(float(v))
    except (TypeError, ValueError):
        return None


def mesmo_jogo(casa1, fora1, casa2, fora2):
    return (
        norm(casa1) == norm(casa2)
        and norm(fora1) == norm(fora2)
    ) or (
        norm(casa1) == norm(fora2)
        and norm(fora1) == norm(casa2)
    )


def main():
    if not FICHEIRO.exists():
        raise FileNotFoundError(
            f"Não encontrei {FICHEIRO}"
        )

    backup = FICHEIRO.with_name(
        f"{FICHEIRO.stem}"
        f"_backup_antes_aplicar_resultados_"
        f"{datetime.now().strftime('%Y%m%d_%H%M%S')}"
        f"{FICHEIRO.suffix}"
    )

    shutil.copy2(FICHEIRO, backup)

    print(f"Backup criado: {backup.name}")

    wb = load_workbook(FICHEIRO)

    ws_atualizacao = wb[SHEET_ATUALIZACAO]
    ws_calendario = wb[SHEET_CALENDARIO]

    resultados_validos = []

    for r in range(2, ws_atualizacao.max_row + 1):
        casa = ws_atualizacao.cell(
            r,
            ATU_CASA_COL,
        ).value

        fora = ws_atualizacao.cell(
            r,
            ATU_FORA_COL,
        ).value

        resultado_casa = numero(
            ws_atualizacao.cell(
                r,
                ATU_RESULTADO_CASA_COL,
            ).value
        )

        resultado_fora = numero(
            ws_atualizacao.cell(
                r,
                ATU_RESULTADO_FORA_COL,
            ).value
        )

        fase = ws_atualizacao.cell(
            r,
            ATU_FASE_COL,
        ).value

        if not casa or not fora:
            continue

        if (
            resultado_casa is None
            or resultado_fora is None
        ):
            continue

        resultados_validos.append(
            {
                "linha": r,
                "fase": fase,
                "casa": casa,
                "fora": fora,
                "resultado_casa": resultado_casa,
                "resultado_fora": resultado_fora,
            }
        )

    print(
        "Resultados válidos encontrados no separador atualização:",
        len(resultados_validos),
    )

    aplicados = 0
    ja_iguais = 0
    nao_encontrados = []

    for resultado in resultados_validos:
        encontrado = False

        for r in range(2, ws_calendario.max_row + 1):
            cal_casa = ws_calendario.cell(
                r,
                CAL_CASA_COL,
            ).value

            cal_fora = ws_calendario.cell(
                r,
                CAL_FORA_COL,
            ).value

            if not mesmo_jogo(
                cal_casa,
                cal_fora,
                resultado["casa"],
                resultado["fora"],
            ):
                continue

            encontrado = True

            # Respeitar orientação casa/fora
            if (
                norm(cal_casa)
                == norm(resultado["casa"])
                and norm(cal_fora)
                == norm(resultado["fora"])
            ):
                novo_casa = resultado["resultado_casa"]
                novo_fora = resultado["resultado_fora"]
            else:
                novo_casa = resultado["resultado_fora"]
                novo_fora = resultado["resultado_casa"]

            atual_casa = numero(
                ws_calendario.cell(
                    r,
                    CAL_REAL_CASA_COL,
                ).value
            )

            atual_fora = numero(
                ws_calendario.cell(
                    r,
                    CAL_REAL_FORA_COL,
                ).value
            )

            if (
                atual_casa == novo_casa
                and atual_fora == novo_fora
            ):
                ja_iguais += 1
            else:
                ws_calendario.cell(
                    r,
                    CAL_REAL_CASA_COL,
                ).value = novo_casa

                ws_calendario.cell(
                    r,
                    CAL_REAL_FORA_COL,
                ).value = novo_fora

                aplicados += 1

                print(
                    f"APLICADO — Calendário linha {r}: "
                    f"{cal_casa} {novo_casa}-{novo_fora} {cal_fora}"
                )

            break

        if not encontrado:
            nao_encontrados.append(
                (
                    resultado["fase"],
                    resultado["casa"],
                    resultado["fora"],
                    resultado["resultado_casa"],
                    resultado["resultado_fora"],
                )
            )

    wb.save(FICHEIRO)
    wb.close()

    print("")
    print("==============================================")
    print("RESUMO")
    print("==============================================")
    print(f"Resultados novos aplicados: {aplicados}")
    print(f"Resultados já iguais: {ja_iguais}")
    print(f"Jogos não encontrados no Calendário: {len(nao_encontrados)}")

    if nao_encontrados:
        print("")
        print("NÃO ENCONTRADOS:")

        for fase, casa, fora, rc, rf in nao_encontrados:
            print(
                f"- {fase} | "
                f"{casa} {rc}-{rf} {fora}"
            )

    print("")
    print("Nenhum prognóstico foi alterado.")


if __name__ == "__main__":
    main()
