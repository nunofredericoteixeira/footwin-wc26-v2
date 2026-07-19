from openpyxl import load_workbook
from pathlib import Path

EXCEL_FILE = Path(__file__).resolve().parent / "mundial_2026_db.xlsx"
SHEET = "Calendário"

def norm(v):
    return str(v or "").strip().lower()

def vencedor(casa, fora, gc, gf):
    if gc is None or gf is None:
        return None
    if int(gc) > int(gf):
        return casa
    if int(gf) > int(gc):
        return fora
    return None

def vencido(casa, fora, gc, gf):
    if gc is None or gf is None:
        return None
    if int(gc) > int(gf):
        return fora
    if int(gf) > int(gc):
        return casa
    return None

def jogo_existe(ws, fase, equipa1, equipa2):
    for row in range(2, ws.max_row + 1):
        f = ws.cell(row, 2).value
        c = ws.cell(row, 3).value
        a = ws.cell(row, 4).value

        if norm(f) != norm(fase):
            continue

        if {norm(c), norm(a)} == {norm(equipa1), norm(equipa2)}:
            return True

    return False

def append_jogo(ws, data, fase, casa, fora, hora):
    nova = ws.max_row + 1

    ws.cell(nova, 1).value = data
    ws.cell(nova, 2).value = fase
    ws.cell(nova, 3).value = casa
    ws.cell(nova, 4).value = fora

    # Resultado real vazio
    ws.cell(nova, 5).value = None
    ws.cell(nova, 6).value = None

    # Prognóstico vazio
    ws.cell(nova, 7).value = None
    ws.cell(nova, 8).value = None
    ws.cell(nova, 9).value = None

    # Tentar colocar hora na mesma coluna usada pelos scripts anteriores.
    # Se houver uma coluna com cabeçalho "Hora", "hora" ou semelhante, usa essa.
    hora_col = None
    for col in range(1, ws.max_column + 1):
        header = norm(ws.cell(1, col).value)
        if header in {"hora", "time", "horário", "horario"}:
            hora_col = col
            break

    if hora_col:
        ws.cell(nova, hora_col).value = hora

    print(f"CRIADO — linha {nova}: {data} {fase} {casa} vs {fora} {hora}")

def main():
    wb = load_workbook(EXCEL_FILE)
    ws = wb[SHEET]

    sfs = []
    for row in range(2, ws.max_row + 1):
        if norm(ws.cell(row, 2).value) == "sf":
            casa = ws.cell(row, 3).value
            fora = ws.cell(row, 4).value
            gc = ws.cell(row, 5).value
            gf = ws.cell(row, 6).value

            win = vencedor(casa, fora, gc, gf)
            lose = vencido(casa, fora, gc, gf)

            if win and lose:
                sfs.append({
                    "row": row,
                    "casa": casa,
                    "fora": fora,
                    "gc": gc,
                    "gf": gf,
                    "winner": win,
                    "loser": lose,
                })

    if len(sfs) < 2:
        print("Ainda não há duas meias-finais com resultado real.")
        print(f"Meias-finais resolvidas encontradas: {len(sfs)}")
        return

    sf1, sf2 = sfs[0], sfs[1]

    terceiro_casa = sf1["loser"]
    terceiro_fora = sf2["loser"]

    final_casa = sf1["winner"]
    final_fora = sf2["winner"]

    print("Meias-finais resolvidas:")
    for sf in sfs:
        print(f"- linha {sf['row']}: {sf['casa']} {sf['gc']}-{sf['gf']} {sf['fora']} | vencedor: {sf['winner']} | vencido: {sf['loser']}")

    if jogo_existe(ws, "3P", terceiro_casa, terceiro_fora):
        print(f"3.º/4.º lugar já existe: {terceiro_casa} vs {terceiro_fora}")
    else:
        append_jogo(ws, "2026-07-18", "3P", terceiro_casa, terceiro_fora, "22:00")

    if jogo_existe(ws, "FINAL", final_casa, final_fora):
        print(f"Final já existe: {final_casa} vs {final_fora}")
    else:
        append_jogo(ws, "2026-07-19", "FINAL", final_casa, final_fora, "20:00")

    wb.save(EXCEL_FILE)
    print("OK: ficheiro guardado.")

if __name__ == "__main__":
    main()
